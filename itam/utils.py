import csv
import io
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage

from flask import Response, g, request
from markupsafe import Markup

from .models import ActivityLog, Notification, Setting, db

# ------------------------------------------------------------------ settings

DEFAULT_SETTINGS = {
    "app_name": "Mada Asset Management System (AMS)",
    "qr_prefix": "",
    "smtp_host": "", "smtp_port": "587", "smtp_user": "", "smtp_password": "",
    "smtp_from": "", "email_enabled": "0",
    "backup_auto": "1", "audit_retention_days": "365",
    "custom_asset_fields": "",
    "checkout_days": "30",
    "warranty_alert_days": "90", "license_alert_days": "90",
    # Printed asset label, in millimetres. The default is the 6 x 3 inch label
    # the app has always produced; set it to match the stock in your printer.
    "label_width_mm": "152.4", "label_height_mm": "76.2",
    # Printed on every label. This is the organisation that owns the asset,
    # which is not the same thing as the name of this software.
    "label_printer": "",
    "label_org": "Mada International Academy",
    # Self-update. Only ever contacts the GitHub Releases API; a failed check
    # is silent. update_token is only needed while the repository is private.
    "update_auto": "1",
    "update_repo": "laithyahya2022-code/Asset-Managment-Software",
    "update_token": "",
    # Set by the updater, not by hand: the version waiting to be applied.
    "update_pending": "",
}

# Ready-made sizes offered on the Settings screen. Roll printers such as the
# Xprinter XP-series take the smaller ones.
LABEL_PRESETS = [
    ("162.6x76.2", "6.4 × 3 in  (163 × 76 mm)"),
    ("152.4x76.2", "6 × 3 in  (152 × 76 mm)"),
    ("101.6x152.4", "4 × 6 in  (102 × 152 mm) shipping roll"),
    ("101.6x50.8", "4 × 2 in  (102 × 51 mm)"),
    ("70x38", "70 × 38 mm"),
    ("55x38", "55 × 38 mm  thermal sticker"),
    ("50x30", "50 × 30 mm  asset tag"),
    ("40x20", "40 × 20 mm  small asset tag"),
]


def label_size_mm():
    """(width, height) of the printed label in mm, clamped to sane values."""
    def read(key, fallback):
        try:
            value = float(get_setting(key))
        except (TypeError, ValueError):
            return fallback
        # Below ~15mm nothing legible fits; above 300mm is past any roll printer.
        return min(max(value, 15.0), 300.0)

    return (read("label_width_mm", 152.4), read("label_height_mm", 76.2))


def get_setting(key):
    row = db.session.get(Setting, key)
    return row.value if row and row.value is not None else DEFAULT_SETTINGS.get(key, "")


def set_setting(key, value):
    row = db.session.get(Setting, key)
    if row:
        row.value = value
    else:
        db.session.add(Setting(key=key, value=value))


def app_name_parts():
    """Split the application name into a short badge and the full title.

    "Mada Asset Management System (AMS)" -> ("AMS", "Mada Asset Management System")
    so the sign-in screen can lead with the abbreviation. A name without a
    trailing "(…)" is used as-is, with no subtitle.
    """
    name = get_setting("app_name").strip()
    match = re.fullmatch(r"(.+?)\s*\(([^()]+)\)", name)
    if match:
        return match.group(2).strip(), match.group(1).strip()
    return name, ""


def label_layout(width_mm=None, height_mm=None, org_name=None):
    """Geometry for the printed asset label, in millimetres.

    The design is a header band (organisation name + asset-tag chip), then
    the detail rows (branch / dept / serial), then a full-width Code 128
    barcode with the tag beneath it. One layout has to cover everything from
    a 40x20mm equipment tag to a 152x76mm sheet, so nothing is fixed: sizes
    derive from the sticker and are squeezed until they fit.
    """
    if width_mm is None or height_mm is None:
        width_mm, height_mm = label_size_mm()
    if org_name is None:
        org_name = get_setting("label_org")
    short = min(width_mm, height_mm)
    portrait = height_mm > width_mm
    pad = max(1.2, short * 0.055)

    header_h = min(max(height_mm * 0.18, 5.5), 13.0)
    fs_org = max(header_h * 0.38, 2.2)
    fs_tag = max(header_h * 0.32, 2.0)
    # The organisation name must survive on one line: shrink it when the
    # header is too narrow for its length (the chip takes ~30% of the width).
    per_line = max(1, int((width_mm * 0.62 - 2 * pad) / (fs_org * 0.52)))
    org_len = len(org_name or "")
    if org_len > per_line:
        fs_org *= max(0.6, per_line / org_len)

    bc_h = min(max(height_mm * 0.20, 4.5), 16.0)
    fs_bctag = max(short * 0.05, 2.0)

    show_org = short >= 24
    if short < 24:
        fields = []
    elif short < 30:
        fields = ["serial"]
    elif short < 34:
        fields = ["branch", "serial"]
    else:
        fields = ["branch", "department", "serial"]

    fs_label = max(short * 0.046, 1.9)
    fs_value = max(short * 0.066, 2.4)
    avail = height_mm - header_h - bc_h - fs_bctag * 1.5 - 2 * pad
    needed = len(fields) * fs_value * 1.7
    if needed > avail > 0:
        squeeze = avail / needed
        fs_label = max(fs_label * squeeze, 1.1)
        fs_value = max(fs_value * squeeze, 1.6)

    return {
        "width": round(width_mm, 2), "height": round(height_mm, 2),
        "portrait": portrait, "pad": round(pad, 2),
        "header_h": round(header_h, 2), "bc_h": round(bc_h, 2),
        "fs_bctag": round(fs_bctag, 2),
        "qr": 0, "gap": round(pad * 0.8, 2),
        "show_org": show_org, "fields": fields, "full_fields": [],
        "org_lines": 1, "columns": 1,
        "fs_org": round(fs_org, 2), "fs_tag": round(fs_tag, 2),
        "fs_label": round(fs_label, 2), "fs_value": round(fs_value, 2),
    }


def custom_field_names():
    return [f.strip() for f in get_setting("custom_asset_fields").split(",") if f.strip()]

# ------------------------------------------------------------------ auditing


def log_activity(action, entity_type=None, entity_id=None, details=None):
    db.session.add(ActivityLog(
        user_id=g.user.id if getattr(g, "user", None) else None,
        action=action, entity_type=entity_type, entity_id=entity_id,
        details=details, ip=request.remote_addr if request else None,
    ))

# -------------------------------------------------------------- notifications


def notify(kind, message, link=None, dedupe_key=None):
    if dedupe_key and db.session.scalar(
            db.select(Notification).where(Notification.dedupe_key == dedupe_key)):
        return
    db.session.add(Notification(kind=kind, message=message, link=link,
                                dedupe_key=dedupe_key))
    if get_setting("email_enabled") == "1":
        send_email(f"[AMS] {kind}", message)


def send_email(subject, body, to=None):
    host = get_setting("smtp_host")
    if not host:
        return False
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = get_setting("smtp_from") or get_setting("smtp_user")
        msg["To"] = to or get_setting("smtp_from") or get_setting("smtp_user")
        msg.set_content(body)
        with smtplib.SMTP(host, int(get_setting("smtp_port") or 587), timeout=10) as s:
            s.starttls()
            if get_setting("smtp_user"):
                s.login(get_setting("smtp_user"), get_setting("smtp_password"))
            s.send_message(msg)
        return True
    except Exception:
        return False

# ------------------------------------------------------------------ CSV


def csv_response(headers, rows, filename):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(["" if v is None else v for v in r])
    return Response(
        "﻿" + buf.getvalue(),  # BOM so Excel opens UTF-8 (Arabic) correctly
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def xlsx_response(headers, rows, filename, sheet="Sheet1"):
    """Build a real .xlsx download from headers + rows."""
    from flask import send_file
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(list(headers))
    for r in rows:
        ws.append([("" if v is None else
                    (v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else
                     (float(v) if hasattr(v, "is_integer") or _is_decimal(v) else v)))
                   for v in r])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


def _is_decimal(v):
    import decimal
    return isinstance(v, decimal.Decimal)


def read_table(file_storage):
    """Read an uploaded .xlsx or .csv into (headers, list-of-dict-rows)."""
    name = (file_storage.filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        data = [[("" if c is None else
                  (c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else c))
                 for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raw = file_storage.read().decode("utf-8-sig", errors="replace")
        data = [row for row in csv.reader(io.StringIO(raw))]
    if not data:
        return [], []
    headers = [str(h or "").strip().lower() for h in data[0]]
    rows = []
    for line in data[1:]:
        if not any(str(c).strip() for c in line):
            continue
        rows.append({headers[i]: (str(line[i]).strip() if i < len(line) else "")
                     for i in range(len(headers))})
    return headers, rows

# ------------------------------------------------------------------ QR / barcode


#: Code 128 bar/space widths, indexed by symbol value 0-106 (106 = stop).
#: Every entry sums to 11 modules except the 13-module stop pattern; the
#: test suite checks those invariants so a typo here cannot slip through.
_C128 = (
    "212222 222122 222221 121223 121322 131222 122213 122312 132212 221213 "
    "221312 231212 112232 122132 122231 113222 123122 123221 223211 221132 "
    "221231 213212 223112 312131 311222 321122 321221 312212 322112 322211 "
    "212123 212321 232121 111323 131123 131321 112313 132113 132311 211313 "
    "231113 231311 112133 112331 132131 113123 113321 133121 313121 211331 "
    "231131 213113 213311 213131 311123 311321 331121 312113 312311 332111 "
    "314111 221411 431111 111224 111422 121124 121421 141122 141221 112214 "
    "112412 122114 122411 142112 142211 241211 221114 413111 241112 134111 "
    "111242 121142 121241 114212 124112 124211 411212 421112 421211 212141 "
    "214121 412121 111143 111341 131141 114113 114311 411113 411311 113141 "
    "114131 311141 411131 211412 211214 211232 2331112"
).split()


def _code128_values(data):
    """Symbol values for `data` in code set B, with start/checksum/stop."""
    vals = [104]                                     # start B
    for ch in str(data):
        v = ord(ch) - 32
        vals.append(v if 0 <= v <= 95 else 31)       # unencodable -> "?"
    checksum = (vals[0] + sum(i * v for i, v in enumerate(vals[1:], 1))) % 103
    return vals + [checksum, 106]


def code128_svg(data):
    """`data` as a Code 128 barcode SVG, for laser scanners that can't
    read QR codes. Stretches to whatever box the label gives it."""
    x, rects = 10.0, []                              # 10-module quiet zone
    for value in _code128_values(data):
        for i, width in enumerate(_C128[value]):
            if i % 2 == 0:                           # even positions are bars
                rects.append(f'<rect x="{x}" y="0" width="{width}" height="10"/>')
            x += int(width)
    return Markup(
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {x + 10} 10" '
        f'preserveAspectRatio="none" shape-rendering="crispEdges" fill="#111">'
        + "".join(rects) + "</svg>")


def qr_svg(data):
    import qrcode
    import qrcode.image.svg
    img = qrcode.make(data, image_factory=qrcode.image.svg.SvgPathImage,
                      box_size=12, border=2)
    return img.to_string().decode()


def barcode_svg(data):
    import barcode
    from barcode.writer import SVGWriter
    buf = io.BytesIO()
    barcode.get("code128", data, writer=SVGWriter()).write(
        buf, options={"module_height": 10, "font_size": 8, "text_distance": 3})
    return buf.getvalue().decode()

# ------------------------------------------------------------------ SVG charts

PALETTE = ["#b8c95e", "#7fa6f2", "#e0a94f", "#b3a1e8", "#6ec4bc",
           "#d05574", "#8a929e", "#4a4f57"]


def bar_chart(pairs, color="#b8c95e", height=210):
    """pairs: [(label, value)] -> responsive SVG bar chart."""
    if not pairs:
        return Markup("<p class='empty'>No data.</p>")
    n = len(pairs)
    bw, gap, pad_l, pad_b = 40, 22, 34, 46
    width = pad_l + n * (bw + gap) + 10
    top = max(v for _, v in pairs) or 1
    bars = []
    for i, (label, v) in enumerate(pairs):
        h = round((v / top) * (height - pad_b - 24), 1)
        x = pad_l + i * (bw + gap)
        y = height - pad_b - h
        short = label if len(label) <= 9 else label[:8] + "…"
        bars.append(
            f'<rect x="{x}" y="{y}" width="{bw}" height="{h}" rx="4" fill="{color}">'
            f'<title>{label}: {v}</title></rect>'
            f'<text x="{x + bw / 2}" y="{y - 6}" text-anchor="middle" class="cv">{v}</text>'
            f'<text x="{x + bw / 2}" y="{height - pad_b + 14}" text-anchor="middle" '
            f'class="cl" transform="rotate(28 {x + bw / 2} {height - pad_b + 14})">{short}</text>'
        )
    grid = "".join(
        f'<line x1="{pad_l - 6}" y1="{height - pad_b - f * (height - pad_b - 24)}" '
        f'x2="{width}" y2="{height - pad_b - f * (height - pad_b - 24)}" class="cg"/>'
        for f in (0, 0.5, 1))
    return Markup(
        f'<svg viewBox="0 0 {width} {height}" class="chart" role="img" '
        f'preserveAspectRatio="xMinYMid meet">{grid}{"".join(bars)}</svg>')


def donut_chart(pairs, size=190):
    """pairs: [(label, value, color)] -> donut with legend."""
    total = sum(v for _, v, _ in pairs)
    if not total:
        return Markup("<p class='empty'>No data.</p>")
    import math
    cx = cy = size / 2
    r, stroke = size / 2 - 16, 26
    circ = 2 * math.pi * r
    offset, segs = 0.0, []
    for label, v, color in pairs:
        frac = v / total
        segs.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{stroke}" stroke-dasharray="{frac * circ} {circ}" '
            f'stroke-dashoffset="{-offset * circ}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{label}: {v}</title></circle>')
        offset += frac
    legend = "".join(
        f'<span class="lg"><i style="background:{c}"></i>{label} ({v})</span>'
        for label, v, c in pairs if v)
    return Markup(
        f'<div class="donut-wrap"><svg viewBox="0 0 {size} {size}" class="chart donut" role="img">'
        f'{"".join(segs)}'
        f'<text x="{cx}" y="{cy - 2}" text-anchor="middle" class="dv">{total}</text>'
        f'<text x="{cx}" y="{cy + 16}" text-anchor="middle" class="dl">assets</text>'
        f'</svg><div class="legend">{legend}</div></div>')


def line_chart(pairs, color="#b8c95e", height=200, money=False):
    """pairs: [(label, value)] month series -> SVG line/area chart."""
    if not pairs:
        return Markup("<p class='empty'>No data.</p>")
    n = len(pairs)
    pad_l, pad_b, pad_t = 40, 30, 14
    width = max(420, n * 52 + pad_l)
    top = max(v for _, v in pairs) or 1
    step = (width - pad_l - 14) / max(1, n - 1)
    pts = []
    for i, (_, v) in enumerate(pairs):
        x = pad_l + i * step
        y = pad_t + (1 - v / top) * (height - pad_b - pad_t)
        pts.append((round(x, 1), round(y, 1)))
    poly = " ".join(f"{x},{y}" for x, y in pts)
    area = f"{pad_l},{height - pad_b} " + poly + f" {pts[-1][0]},{height - pad_b}"
    labels = "".join(
        f'<text x="{pts[i][0]}" y="{height - 8}" text-anchor="middle" class="cl">{label}</text>'
        for i, (label, _) in enumerate(pairs))
    dots = "".join(f'<circle cx="{x}" cy="{y}" r="3.5" fill="{color}"/>' for x, y in pts)
    fmt = (lambda v: f"{v:,.0f}") if money else (lambda v: v)
    vals = "".join(
        f'<text x="{pts[i][0]}" y="{pts[i][1] - 8}" text-anchor="middle" class="cv">{fmt(v)}</text>'
        for i, (_, v) in enumerate(pairs))
    return Markup(
        f'<svg viewBox="0 0 {width} {height}" class="chart" role="img">'
        f'<polygon points="{area}" fill="{color}" opacity="0.12"/>'
        f'<polyline points="{poly}" fill="none" stroke="{color}" stroke-width="2.5"/>'
        f'{dots}{vals}{labels}</svg>')

# ------------------------------------------------------------------ misc


def parse_date(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def month_key(d):
    return d.strftime("%b %y")


# --------------------------------------------------------------- location levels

#: Which Location kind backs each level of the asset form.
LEVEL_KINDS = {"branch": "Branch", "building": "Building",
               "floor": "Floor", "room": "Room"}


def level_values(level):
    """Every name in use for one level of the hierarchy.

    The asset form used to offer -- and validate against -- lists compiled
    into the app, so a school could not record a fourth campus or a new
    annexe: the value was silently dropped on save. This reads what actually
    exists, plus the standard names as a starting point.
    """
    from .models import (BRANCHES, BUILDINGS, FLOORS, PLACES, Asset, Location,
                         db)

    kind = LEVEL_KINDS.get(level)
    standard = {"branch": BRANCHES, "building": BUILDINGS,
                "floor": FLOORS, "room": PLACES}.get(level, [])
    names = set(standard)
    if kind:
        names |= {n for (n,) in db.session.execute(
            db.select(Location.name).where(Location.kind == kind).distinct())}
        if kind == "Room":       # storage areas are rooms as far as an asset cares
            names |= {n for (n,) in db.session.execute(
                db.select(Location.name).where(Location.kind == "Storage Area").distinct())}
    # Whatever the register itself already says, so an import is never lost.
    column = {"branch": Asset.branch, "building": Asset.building,
              "floor": Asset.floor, "room": Asset.location_name}.get(level)
    if column is not None:
        names |= {n for (n,) in db.session.execute(
            db.select(column).where(column.isnot(None), column != "").distinct())}
    return sorted(n for n in names if n)
