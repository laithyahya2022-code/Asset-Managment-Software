import csv
import io
import json
import os
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal

from flask import (Blueprint, Response, current_app, flash, g, jsonify,
                   redirect, render_template, request, send_from_directory,
                   url_for)
from sqlalchemy import or_
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.utils import secure_filename

from ..models import (ASSET_CONDITIONS, ASSET_STATUSES, BRANCHES, BUILDINGS,
                      BLOCKED_CHECKOUT_STATUSES, FLOORS, OPERATING_SYSTEMS,
                      PLACES, ActivityLog, Asset,
                      Assignment, Category, Department, Employee, Location,
                      Reservation, Transfer, Vendor, db)
from ..security import has_perm, login_required, perm_required
from ..utils import (barcode_svg, csv_response, custom_field_names,
                     get_setting, label_layout, level_values, log_activity,
                     parse_date, qr_svg)

bp = Blueprint("assets", __name__, url_prefix="/assets")


# --------------------------------------------------------------- undo support
# A few actions are reversible by an admin from the Activity log. We keep just
# enough on the ActivityLog row (undo_data JSON) to reverse them:
#   - an import records the ids it created   -> undo deletes them
#   - a delete records a snapshot of each asset -> undo re-creates them
# Undo runs once (undone_at guards it) and asks for confirmation first.
_SNAPSHOT_SKIP = {"id", "created_at", "updated_at"}


def _asset_snapshot(a):
    """A JSON-safe copy of an asset's columns, enough to re-create it."""
    snap = {}
    for col in Asset.__table__.columns:
        if col.name in _SNAPSHOT_SKIP:
            continue
        v = getattr(a, col.name)
        if isinstance(v, (date, datetime)):
            v = v.isoformat()
        elif isinstance(v, Decimal):
            v = float(v)
        snap[col.name] = v
    return snap


def _restore_asset(snap):
    """Re-create an asset from a snapshot (new id; a dangling parent is dropped)."""
    kwargs = {}
    for col in Asset.__table__.columns:
        if col.name in _SNAPSHOT_SKIP or col.name not in snap:
            continue
        v = snap[col.name]
        if isinstance(v, str) and v:
            kind = str(col.type).lower()
            try:
                if "datetime" in kind:
                    v = datetime.fromisoformat(v)
                elif "date" in kind:
                    v = date.fromisoformat(v)
            except ValueError:
                pass
        kwargs[col.name] = v
    a = Asset(**kwargs)
    a.parent_id = None      # the parent may not have been restored
    db.session.add(a)
    return a


def undo_activity(log):
    """Reverse a reversible activity. Returns (ok, message)."""
    if not log.undo_data or log.undone_at:
        return False, "This action can't be undone."
    data = json.loads(log.undo_data)
    op = data.get("op")
    if op == "delete":                       # undo an import: remove what it made
        ids = data.get("asset_ids", [])
        assets = db.session.scalars(
            db.select(Asset).where(Asset.id.in_(ids))).all()
        for a in assets:
            db.session.delete(a)
        msg = f"Undone — {len(assets)} imported assets removed."
    elif op == "restore":                    # undo a delete: bring the assets back
        snaps = data.get("assets", [])
        for snap in snaps:
            _restore_asset(snap)
        msg = f"Undone — {len(snaps)} assets restored."
    else:
        return False, "This action can't be undone."
    log.undone_at = datetime.utcnow()
    log_activity("undone", log.entity_type, None,
                 f"Undid: {log.action} — {log.details or ''}".strip())
    db.session.commit()
    _refresh_locations()
    return True, msg


@bp.post("/undo/<int:log_id>")
@perm_required("assets.manage")
def undo(log_id):
    log = db.get_or_404(ActivityLog, log_id)
    ok, msg = undo_activity(log)
    flash(msg, "success" if ok else "error")
    return redirect(request.referrer or url_for("main.activity"))

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}
ALLOWED_EXTS = IMAGE_EXTS | {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".txt", ".csv"}


def _lookups():
    return dict(
        categories=db.session.scalars(db.select(Category).order_by(Category.name)).all(),
        departments=db.session.scalars(db.select(Department).order_by(Department.name)).all(),
        locations=db.session.scalars(db.select(Location).order_by(Location.name)).all(),
        vendors=db.session.scalars(db.select(Vendor).order_by(Vendor.name)).all(),
        employees=db.session.scalars(db.select(Employee).where(Employee.active)
                                     .order_by(Employee.name)).all(),
        statuses=ASSET_STATUSES, conditions=ASSET_CONDITIONS,
        # Every level comes from the Locations screen plus whatever the
        # register already uses, so a branch or building added there is
        # immediately available here.
        branches=level_values("branch"), buildings=level_values("building"),
        floors=level_values("floor"), rooms=level_values("room"),
        places=level_values("room"),
        operating_systems=OPERATING_SYSTEMS, today_iso=date.today().isoformat(),
        custom_names=custom_field_names(),
    )


def next_tag(category):
    """The next Asset ID for a category, in the format the register already uses.

    This only ever looked for "PREFIX-999999". A school whose asset IDs are
    written PC00010, with no separator, matched nothing: the count restarted
    at 1 *and* switched format, so picking a category on a register of a
    thousand PCs suggested "PC-000001". Read the separator and the digit width
    off the tags that exist and carry on from the highest.
    """
    import re
    from collections import Counter

    prefix = category.tag_prefix if category else "AST"
    top, shapes = 0, Counter()
    pattern = re.compile(rf"^{re.escape(prefix)}([-_/ ]?)(\d+)$", re.IGNORECASE)
    for (tag,) in db.session.execute(
            db.select(Asset.tag).where(Asset.tag.like(f"{prefix}%"))).all():
        match = pattern.match(tag or "")
        if match:
            shapes[(match.group(1), len(match.group(2)))] += 1
            top = max(top, int(match.group(2)))
    # Follow the majority shape; a register with none falls back to PC-000001.
    separator, width = shapes.most_common(1)[0][0] if shapes else ("-", 6)
    return f"{prefix}{separator}{top + 1:0{width}d}"


def _next_tag_map():
    """{category_id: next auto-generated Asset ID} for live form preview."""
    return {c.id: next_tag(c) for c in
            db.session.scalars(db.select(Category)).all()}


def _from_form(a, form):
    a.name = form["name"].strip()
    a.category_id = int(form["category_id"]) if form.get("category_id") else None
    tag = form.get("tag", "").strip()
    a.tag = tag or next_tag(db.session.get(Category, a.category_id)
                            if a.category_id else None)
    a.serial = form.get("serial", "").strip() or None
    a.manufacturer = form.get("manufacturer", "").strip() or None
    a.model = form.get("model", "").strip() or None
    if form.get("status") in ASSET_STATUSES:
        a.status = form["status"]
    if form.get("condition") in ASSET_CONDITIONS:
        a.condition = form["condition"]
    # Only write fields the form actually submitted. "OS version", "Type" and
    # "Device name" have all been taken off the form at various points, and
    # blanking them here would erase values that came in from an import.
    for field in ("asset_type", "os_name", "os_version", "cpu", "ram",
                  "storage", "gpu", "hostname", "mac_address", "ip_address",
                  "invoice_number"):
        if field in form:
            setattr(a, field, form.get(field, "").strip() or None)
    # Validate against the levels that actually exist, not a list compiled
    # into the app: a school with a fourth campus used to have the value
    # silently thrown away on save.
    for level in ("branch", "building", "floor"):
        value = form.get(level, "").strip()
        setattr(a, level, value if value in level_values(level) else None)
    a.location_name = form.get("location_name", "").strip() or None
    if "department_id" in form:
        dep = form.get("department_id")
        a.department_id = int(dep) if dep else None
    # "Updated by" is shown on the asset form but no longer editable there, so
    # the field isn't submitted. Stamp whoever saved the record instead; the
    # importer and the API still pass their own value and keep it.
    if "updated_by" in form:
        a.updated_by = form.get("updated_by", "").strip() or None
    elif getattr(g, "user", None) is not None:
        a.updated_by = g.user.name
    a.vendor_id = int(form["vendor_id"]) if form.get("vendor_id") else None
    a.purchase_date = parse_date(form.get("purchase_date"))
    a.purchase_cost = form.get("purchase_cost") or None
    a.depreciation_years = int(form.get("depreciation_years") or 5)
    a.warranty_expiry = parse_date(form.get("warranty_expiry"))
    a.notes = form.get("notes", "").strip() or None
    custom = {name: form.get(f"custom_{i}", "").strip()
              for i, name in enumerate(custom_field_names())}
    a.custom_fields = json.dumps({k: v for k, v in custom.items() if v})
    # The asset form no longer offers "Part of (parent asset)". Only touch the
    # link when a form actually submits the field, otherwise saving an edit
    # would silently unlink assets that already have a parent.
    if "parent_id" in form:
        parent = form.get("parent_id")
        a.parent_id = int(parent) if parent and (not a.id or int(parent) != a.id) else None


def _refresh_locations():
    """Keep Location rows in step with what the assets say.

    The Locations screen used to do this on page load, but that screen is
    gone: a place is a property of an asset now, not something maintained
    separately. The rows still back the location filter, bulk transfers,
    transfer history and the locations report, so they are refreshed at the
    only moments the data can change -- an asset being saved or imported.
    """
    from .org import sync_locations_from_assets
    try:
        sync_locations_from_assets()
    except Exception:
        db.session.rollback()   # never fail a save over a derived side table


def _apply_assignment(a, form):
    """Handle the 'Assign to' field on the asset form (spec section 15)."""
    emp_id = form.get("assign_employee_id")
    cur = a.current_assignment
    if not emp_id:
        return
    emp_id = int(emp_id)
    if cur and cur.employee_id == emp_id:
        return  # already assigned to this person
    if a.status in BLOCKED_CHECKOUT_STATUSES:
        return
    if cur:  # reassignment: return the current holder first
        cur.returned_at = datetime.utcnow()
    emp = db.session.get(Employee, emp_id)
    if not emp:
        return
    db.session.add(Assignment(asset=a, employee=emp, assigned_by=g.user.id,
                              handled_by=g.user.name))
    a.status = "Checked Out"
    log_activity("checked_out", "asset", a.id, f"{a.tag} → {emp.name}")


#: Rows per page on the asset list. Rendering every row of a 3,000-asset
#: register produced a 1 MB page and one SQL query per asset.
PAGE_SIZE = 50


def _assets_query(args):
    # The list shows category, department, location and the current holder for
    # every row. Without this each row fetched its own, which is where the
    # 3,000-query page load came from.
    stmt = db.select(Asset).options(
        joinedload(Asset.category),
        joinedload(Asset.department),
        joinedload(Asset.location),
        selectinload(Asset.assignments).joinedload(Assignment.employee),
    ).order_by(Asset.tag)
    q = args.get("q", "").strip()
    if q:
        like = f"%{q}%"
        stmt = stmt.where(or_(Asset.tag.ilike(like), Asset.name.ilike(like),
                              Asset.serial.ilike(like), Asset.manufacturer.ilike(like),
                              Asset.model.ilike(like), Asset.asset_type.ilike(like),
                              Asset.hostname.ilike(like), Asset.mac_address.ilike(like),
                              Asset.ip_address.ilike(like), Asset.os_name.ilike(like)))
    if args.get("status"):
        stmt = stmt.where(Asset.status == args["status"])
    # A hand-edited URL or a stale bookmark used to reach int() directly and
    # crash the page with a 500. A filter that isn't a number is no filter.
    for param, column in (("category", Asset.category_id),
                          ("department", Asset.department_id),
                          ("location", Asset.location_id)):
        raw = args.get(param)
        if raw:
            try:
                stmt = stmt.where(column == int(raw))
            except (TypeError, ValueError):
                pass
    if args.get("condition"):
        stmt = stmt.where(Asset.condition == args["condition"])
    if args.get("branch"):
        stmt = stmt.where(Asset.branch == args["branch"])
    if args.get("building"):
        stmt = stmt.where(Asset.building == args["building"])
    if args.get("floor"):
        stmt = stmt.where(Asset.floor == args["floor"])
    return stmt


def _filtered_assets(args):
    """Every matching asset — used by exports and the label sheet."""
    return db.session.scalars(_assets_query(args)).unique().all()


def _assets_page(args):
    """One page of matching assets, plus the paging figures for the template."""
    try:
        page = max(1, int(args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    stmt = _assets_query(args)
    total = db.session.scalar(
        db.select(db.func.count()).select_from(stmt.order_by(None).subquery()))
    pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = min(page, pages)
    rows = db.session.scalars(
        stmt.limit(PAGE_SIZE).offset((page - 1) * PAGE_SIZE)).unique().all()
    return rows, {"page": page, "pages": pages, "total": total,
                  "size": PAGE_SIZE,
                  "first": 0 if not total else (page - 1) * PAGE_SIZE + 1,
                  "last": min(page * PAGE_SIZE, total)}


def _distinct(col):
    """Sorted list of the non-empty values actually present in the data."""
    vals = db.session.scalars(
        db.select(col).where(col.isnot(None), col != "").distinct()).all()
    return sorted(vals)


def _locations_in_use(all_locations, keep_id=None):
    """Locations worth offering as a filter.

    The standard tree is ~600 rows, which made a 39 KB <select> on the busiest
    page in the app. Filtering by a location that holds nothing returns
    nothing, so only list the ones actually in use — plus whichever is
    currently selected, so an active filter never vanishes from its own box.
    """
    used = set(db.session.scalars(
        db.select(Asset.location_id).where(Asset.location_id.isnot(None)).distinct()))
    if keep_id:
        used.add(keep_id)
    return [l for l in all_locations if l.id in used]


@bp.route("/")
@perm_required("assets.view")
def list_():
    from ..models import SavedSearch
    searches = db.session.scalars(db.select(SavedSearch)
                                  .where(SavedSearch.user_id == g.user.id)
                                  .order_by(SavedSearch.name)).all()
    rows, paging = _assets_page(request.args)
    looks = _lookups()
    try:
        active_loc = int(request.args.get("location") or 0)
    except (TypeError, ValueError):
        active_loc = 0
    return render_template("assets/list.html", assets=rows, paging=paging,
                           args=request.args, searches=searches,
                           branch_opts=_distinct(Asset.branch),
                           building_opts=_distinct(Asset.building),
                           floor_opts=_distinct(Asset.floor),
                           filter_locations=_locations_in_use(
                               looks["locations"], active_loc),
                           **looks)


@bp.get("/locations.json")
@perm_required("assets.view")
def locations_json():
    """Feeds the bulk "transfer to location" picker.

    Unlike the filter, a transfer target may well be an empty location, so
    this searches the whole tree rather than only what's in use.
    """
    q = request.args.get("q", "").strip().lower()
    rows = db.session.scalars(db.select(Location).order_by(Location.name)).all()
    hits = [l for l in rows if not q or q in l.path.lower()]
    hits.sort(key=lambda l: l.path)
    return jsonify([{"id": l.id, "path": l.path} for l in hits[:200]])


@bp.post("/searches")
@perm_required("assets.view")
def search_save():
    from ..models import SavedSearch
    name = request.form.get("name", "").strip()
    query = request.form.get("query", "").strip()
    if not name or not query:
        flash("Apply some filters first, then give the search a name.", "error")
    else:
        db.session.add(SavedSearch(user_id=g.user.id, name=name[:80], query=query[:500]))
        db.session.commit()
        flash(f'Search "{name}" saved.', "success")
    return redirect(url_for("assets.list_") + ("?" + query if query else ""))


@bp.post("/searches/<int:search_id>/delete")
@perm_required("assets.view")
def search_delete(search_id):
    from ..models import SavedSearch
    s = db.get_or_404(SavedSearch, search_id)
    if s.user_id == g.user.id:
        db.session.delete(s)
        db.session.commit()
        flash("Saved search removed.", "success")
    return redirect(url_for("assets.list_"))


# The export mirrors the school's own inventory sheet, column for column, so
# a file that went in comes back out arranged the same way (and can be
# re-imported as-is). Two register-only columns ride along at the end.
EXPORT_HEADERS = [
    "Name", "Category", "Asset ID", "Condition", "Status", "Serial No.",
    "Manufacturer", "Model", "Branch", "Building", "Floor", "Room",
    "Department", "Assign to", "Updated By", "Vendor", "Purchase Date",
    "Purchase Cost", "Depreciation (Years)", "Warranty Expiry",
    "Invoice Number", "Operating system", "CPU", "RAM", "Storage",
    "Graphic cards", "MAC Address", "IP address", "Notes",
    "Type", "Hostname"]


def _export_row(a, blank=""):
    return [a.name, a.category.name if a.category else blank, a.tag,
            a.condition, a.status, a.serial or blank, a.manufacturer or blank,
            a.model or blank, a.branch or blank, a.building or blank,
            a.floor or blank,
            a.location_name or (a.location.path if a.location else blank),
            a.department.name if a.department else blank,
            a.assigned_label or blank,
            a.updated_by or blank, a.vendor.name if a.vendor else blank,
            a.purchase_date or blank,
            float(a.purchase_cost) if a.purchase_cost else blank,
            a.depreciation_years or blank, a.warranty_expiry or blank,
            a.invoice_number or blank, a.os_name or blank, a.cpu or blank,
            a.ram or blank, a.storage or blank, a.gpu or blank,
            a.mac_address or blank, a.ip_address or blank, a.notes or blank,
            a.asset_type or blank, a.hostname or blank]


@bp.route("/export.csv")
@perm_required("assets.view")
def export_csv():
    rows = [_export_row(a) for a in _filtered_assets(request.args)]
    return csv_response(EXPORT_HEADERS, rows, "assets.csv")


@bp.route("/new", methods=["GET", "POST"])
@perm_required("assets.manage")
def new():
    source = None
    if request.args.get("clone"):
        try:
            source = db.session.get(Asset, int(request.args["clone"]))
        except (TypeError, ValueError):
            source = None       # a junk ?clone= is not worth a 500
    if request.method == "POST":
        tag = request.form.get("tag", "").strip()
        if tag and db.session.scalar(db.select(Asset).where(Asset.tag == tag)):
            flash(f'Asset tag "{tag}" already exists.', "error")
        else:
            a = Asset()
            _from_form(a, request.form)
            db.session.add(a)
            db.session.flush()
            log_activity("created", "asset", a.id, f"{a.tag} — {a.name}")
            _apply_assignment(a, request.form)
            db.session.commit()
            _refresh_locations()
            flash(f"Asset {a.tag} created.", "success")
            return redirect(url_for("assets.detail", asset_id=a.id))
    return render_template("assets/form.html", asset=None, source=source,
                           next_tags=_next_tag_map(),
                           qr_prefix=get_setting("qr_prefix"), **_lookups())


@bp.route("/<int:asset_id>")
@perm_required("assets.view")
def detail(asset_id):
    from ..models import ActivityLog, InventoryCheck
    a = db.get_or_404(Asset, asset_id)
    timeline = db.session.scalars(
        db.select(ActivityLog)
        .where(ActivityLog.entity_type == "asset", ActivityLog.entity_id == a.id)
        .order_by(ActivityLog.created_at.desc()).limit(12)).all()
    last_check = db.session.scalars(
        db.select(InventoryCheck).where(InventoryCheck.asset_id == a.id)
        .order_by(InventoryCheck.checked_at.desc()).limit(1)).first()
    return render_template("assets/detail.html", asset=a, today=date.today(),
                           timeline=timeline, last_check=last_check,
                           L=label_layout(),
                           default_due=(date.today() + timedelta(
                               days=int(get_setting("checkout_days") or 30))),
                           image_exts=IMAGE_EXTS, **_lookups())


@bp.route("/<int:asset_id>/edit", methods=["GET", "POST"])
@perm_required("assets.manage")
def edit(asset_id):
    a = db.get_or_404(Asset, asset_id)
    if request.method == "POST":
        tag = request.form["tag"].strip()
        other = db.session.scalar(db.select(Asset).where(Asset.tag == tag))
        if other and other.id != a.id:
            flash(f'Asset tag "{tag}" already exists.', "error")
        else:
            _from_form(a, request.form)
            log_activity("updated", "asset", a.id, a.tag)
            _apply_assignment(a, request.form)
            db.session.commit()
            _refresh_locations()
            flash(f"Asset {a.tag} updated.", "success")
            return redirect(url_for("assets.detail", asset_id=a.id))
    return render_template("assets/form.html", asset=a, source=None,
                           next_tags=_next_tag_map(),
                           qr_prefix=get_setting("qr_prefix"), **_lookups())


@bp.post("/<int:asset_id>/delete")
@perm_required("assets.manage")
def delete(asset_id):
    a = db.get_or_404(Asset, asset_id)
    log_activity("deleted", "asset", a.id, a.tag,
                 undo_data=json.dumps({"op": "restore", "assets": [_asset_snapshot(a)]}))
    db.session.delete(a)
    db.session.commit()
    flash(f"Asset {a.tag} deleted.", "success")
    return redirect(url_for("assets.list_"))


# ----------------------------------------------------------------- categories

@bp.route("/categories", methods=["GET", "POST"])
@perm_required("assets.view")
def categories():
    if request.method == "POST":
        if not has_perm("assets.manage"):
            flash("You do not have permission to do that.", "error")
            return redirect(url_for("assets.categories"))
        name = request.form["name"].strip()
        prefix = request.form.get("prefix", "").strip().upper() or None
        if name and not db.session.scalar(db.select(Category).where(Category.name == name)):
            db.session.add(Category(name=name, prefix=prefix))
            log_activity("category_created", "category", None, name)
            db.session.commit()
            flash(f'Category "{name}" created.', "success")
        else:
            flash("Category name is empty or already exists.", "error")
        return redirect(url_for("assets.categories"))
    rows = db.session.scalars(db.select(Category).order_by(Category.name)).all()
    return render_template("assets/categories.html", rows=rows)


@bp.post("/categories/<int:cat_id>/delete")
@perm_required("assets.manage")
def category_delete(cat_id):
    cat = db.get_or_404(Category, cat_id)
    if cat.assets:
        flash("Category is in use by assets.", "error")
    else:
        db.session.delete(cat)
        db.session.commit()
        flash("Category deleted.", "success")
    return redirect(url_for("assets.categories"))


# ------------------------------------------------------------ lifecycle ops

@bp.post("/<int:asset_id>/checkout")
@perm_required("checkout.manage")
def checkout(asset_id):
    a = db.get_or_404(Asset, asset_id)
    if a.current_assignment or a.status in BLOCKED_CHECKOUT_STATUSES:
        flash("This asset cannot be checked out.", "error")
        return redirect(url_for("assets.detail", asset_id=a.id))
    emp = db.get_or_404(Employee, int(request.form["employee_id"]))
    db.session.add(Assignment(asset=a, employee=emp, assigned_by=g.user.id,
                              due_at=parse_date(request.form.get("due_at")),
                              handled_by=g.user.name,
                              notes=request.form.get("notes", "").strip() or None))
    a.status = "Checked Out"
    a.updated_by = g.user.name
    log_activity("checked_out", "asset", a.id, f"{a.tag} → {emp.name}")
    db.session.commit()
    flash(f"Checked out to {emp.name}.", "success")
    return redirect(url_for("assets.detail", asset_id=a.id))


@bp.post("/<int:asset_id>/checkin")
@perm_required("checkout.manage")
def checkin(asset_id):
    a = db.get_or_404(Asset, asset_id)
    asg = a.current_assignment
    if not asg:
        flash("Asset is not checked out.", "error")
        return redirect(url_for("assets.detail", asset_id=a.id))
    asg.returned_at = datetime.utcnow()
    # return inspection (spec section 16)
    condition = request.form.get("return_condition", "")
    if condition in ASSET_CONDITIONS and condition != a.condition:
        log_activity("condition_changed", "asset", a.id,
                     f"{a.tag}: {a.condition} → {condition} on return")
        a.condition = condition
        asg.return_condition = condition
    asg.return_notes = request.form.get("return_notes", "").strip() or None
    a.status = "Damaged" if condition == "Damaged" else "Available"
    log_activity("checked_in", "asset", a.id, f"{a.tag} ← {asg.employee.name}")
    db.session.commit()
    flash(f"Checked in from {asg.employee.name}.", "success")
    return redirect(url_for("assets.detail", asset_id=a.id))


@bp.post("/<int:asset_id>/transfer")
@perm_required("checkout.manage")
def transfer(asset_id):
    """Change where the asset lives, with the same Branch / Building / Floor /
    Room / Department choices the asset form offers. The old single
    "location_id" form is still honoured for the bulk action."""
    a = db.get_or_404(Asset, asset_id)
    note = request.form.get("notes", "").strip()

    if request.form.get("location_id"):              # legacy single-picker form
        to_id = int(request.form["location_id"])
        db.session.add(Transfer(asset=a, from_location_id=a.location_id,
                                to_location_id=to_id, by_user=g.user.id,
                                notes=note or None))
        a.location_id = to_id
    else:
        def where(asset):
            return " / ".join(p for p in (asset.branch, asset.building,
                                          asset.floor, asset.location_name)
                              if p) or "—"

        before = where(a)
        a.branch = request.form.get("branch", "").strip() or None
        a.building = request.form.get("building", "").strip() or None
        a.floor = request.form.get("floor", "").strip() or None
        a.location_name = request.form.get("location_name", "").strip() or None
        dep = request.form.get("department_id", "").strip()
        a.department_id = int(dep) if dep.isdigit() else None
        after = where(a)
        if before == after and not note:
            flash("Nothing changed.", "error")
            return redirect(url_for("assets.detail", asset_id=a.id))
        db.session.add(Transfer(
            asset=a, by_user=g.user.id,
            notes=f"{before} → {after}" + (f" — {note}" if note else "")))
        _refresh_locations()
    log_activity("transferred", "asset", a.id, a.tag)
    db.session.commit()
    flash("Asset transferred.", "success")
    return redirect(url_for("assets.detail", asset_id=a.id))


@bp.post("/<int:asset_id>/reserve")
@perm_required("checkout.manage")
def reserve(asset_id):
    a = db.get_or_404(Asset, asset_id)
    start = parse_date(request.form["start_date"])
    end = parse_date(request.form["end_date"])
    if not start or not end or end < start:
        flash("Enter a valid reservation period.", "error")
        return redirect(url_for("assets.detail", asset_id=a.id))
    db.session.add(Reservation(asset=a, employee_id=int(request.form["employee_id"]),
                               start_date=start, end_date=end,
                               notes=request.form.get("notes", "").strip() or None))
    log_activity("reserved", "asset", a.id, a.tag)
    db.session.commit()
    flash("Reservation created.", "success")
    return redirect(url_for("assets.detail", asset_id=a.id))


# ------------------------------------------------------------------ files

@bp.post("/<int:asset_id>/files")
@perm_required("assets.manage")
def upload_file(asset_id):
    a = db.get_or_404(Asset, asset_id)
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Choose a file to upload.", "error")
        return redirect(url_for("assets.detail", asset_id=a.id))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTS:
        flash(f"File type {ext} is not allowed.", "error")
        return redirect(url_for("assets.detail", asset_id=a.id))
    stored = f"{uuid.uuid4().hex}{ext}"
    f.save(os.path.join(current_app.config["UPLOAD_FOLDER"], stored))
    from ..models import AssetFile
    db.session.add(AssetFile(asset=a, stored_name=stored,
                             orig_name=secure_filename(f.filename) or f"file{ext}",
                             kind=request.form.get("kind", "document"),
                             uploaded_by=g.user.id))
    log_activity("file_uploaded", "asset", a.id, f.filename)
    db.session.commit()
    flash("File uploaded.", "success")
    return redirect(url_for("assets.detail", asset_id=a.id))


@bp.route("/files/<int:file_id>")
@perm_required("assets.view")
def get_file(file_id):
    from ..models import AssetFile
    af = db.get_or_404(AssetFile, file_id)
    return send_from_directory(current_app.config["UPLOAD_FOLDER"], af.stored_name,
                               download_name=af.orig_name,
                               as_attachment=request.args.get("dl") == "1")


@bp.post("/files/<int:file_id>/delete")
@perm_required("assets.manage")
def delete_file(file_id):
    from ..models import AssetFile
    af = db.get_or_404(AssetFile, file_id)
    asset_id = af.asset_id
    try:
        os.remove(os.path.join(current_app.config["UPLOAD_FOLDER"], af.stored_name))
    except OSError:
        pass
    db.session.delete(af)
    db.session.commit()
    flash("File deleted.", "success")
    return redirect(url_for("assets.detail", asset_id=asset_id))


# ------------------------------------------------------------ QR / barcode

@bp.route("/qr-preview.svg")
@perm_required("assets.view")
def qr_preview():
    data = request.args.get("data", "").strip() or "PREVIEW"
    return Response(qr_svg(get_setting("qr_prefix") + data), mimetype="image/svg+xml")


@bp.get("/barcode-preview")
@perm_required("assets.view")
def barcode_preview():
    """Live Code 128 for the label preview on the asset form."""
    from ..utils import code128_svg
    data = request.args.get("data", "").strip() or "PREVIEW"
    return Response(code128_svg(data), mimetype="image/svg+xml")


@bp.route("/<int:asset_id>/qr.svg")
@perm_required("assets.view")
def qr(asset_id):
    a = db.get_or_404(Asset, asset_id)
    return Response(qr_svg(get_setting("qr_prefix") + a.tag), mimetype="image/svg+xml")


@bp.route("/<int:asset_id>/barcode.svg")
@perm_required("assets.view")
def barcode(asset_id):
    a = db.get_or_404(Asset, asset_id)
    return Response(barcode_svg(a.tag), mimetype="image/svg+xml")


@bp.route("/<int:asset_id>/label")
@perm_required("assets.view")
def label(asset_id):
    a = db.get_or_404(Asset, asset_id)
    bare = request.args.get("bare") == "1"       # embedded preview, no chrome
    L = label_layout()
    html = render_template("assets/label.html", asset=a,
                           app_name=get_setting("label_org"),
                           L=L, bare=bare,
                           auto=not bare and request.args.get("auto") == "1")
    if bare:
        return html
    return _print_or_show(html, f"Label for {a.tag}", tspl_assets=[a], L=L)


def _label_rows(a, L):
    """The (field, value) rows the label design shows for this asset."""
    values = {"branch": a.branch,
              "department": a.department.name if a.department else None,
              "serial": a.serial}
    titles = {"branch": "Branch", "department": "Dept", "serial": "S/N"}
    return [(titles[k], values[k] or "-") for k in L["fields"]]


def _print_or_show(html, what, tspl_assets=None, L=None):
    """Send the label to the configured printer, or fall back to the browser.

    With a printer named in Settings and AMS running as the packaged Windows
    app, the label is sent to the printer in its native TSPL language --
    exact sticker size, gap registration and rotation built in, with no
    browser or driver page pipeline to rotate it. If that fails, the old
    print-via-browser path runs; failing that, the page just opens.
    """
    from .. import printing

    printer = get_setting("label_printer")
    if printing.can_print_directly(printer, request.remote_addr):
        if tspl_assets and L:
            program = "".join(
                printing.label_tspl(L["width"], L["height"],
                                    get_setting("label_org"), a.tag,
                                    _label_rows(a, L), flip=L["rotate"])
                for a in tspl_assets)
            if printing.print_raw(printer, program):
                flash(f"{what} sent to {printer}.", "success")
                return redirect(request.referrer or url_for("assets.list_"))
        if printing.print_html(html, printer):
            flash(f"{what} sent to {printer}.", "success")
            return redirect(request.referrer or url_for("assets.list_"))
    return html


@bp.route("/labels")
@perm_required("assets.view")
def labels():
    ids = request.args.getlist("id", type=int)
    stmt = db.select(Asset).order_by(Asset.tag)
    if ids:
        stmt = stmt.where(Asset.id.in_(ids))
    rows = db.session.scalars(stmt).all()
    L = label_layout()
    html = render_template("assets/labels.html", assets=rows,
                           app_name=get_setting("label_org"), L=L)
    return _print_or_show(html, f"{len(rows)} labels", tspl_assets=rows, L=L)


# ------------------------------------------------------------------ bulk

@bp.post("/bulk")
@perm_required("assets.manage")
def bulk():
    ids = request.form.getlist("id", type=int)
    action = request.form.get("action")
    if not ids:
        flash("Select at least one asset.", "error")
        return redirect(url_for("assets.list_"))
    if action == "labels":
        return redirect(url_for("assets.labels", id=ids))
    assets = db.session.scalars(db.select(Asset).where(Asset.id.in_(ids))).all()
    if action == "delete":
        snaps = [_asset_snapshot(a) for a in assets]
        log_activity("deleted", "asset", None, f"{len(assets)} assets",
                     undo_data=json.dumps({"op": "restore", "assets": snaps}))
        for a in assets:
            db.session.delete(a)
        flash(f"{len(assets)} assets deleted.", "success")
    elif action and action.startswith("status:"):
        status = action.split(":", 1)[1]
        if status in ASSET_STATUSES:
            for a in assets:
                a.status = status
                log_activity("updated", "asset", a.id, f"{a.tag} status → {status}")
            flash(f"{len(assets)} assets set to {status}.", "success")
    elif action == "assign" and request.form.get("employee_id"):
        emp = db.get_or_404(Employee, int(request.form["employee_id"]))
        done = skipped = 0
        for a in assets:
            if a.current_assignment or a.status in BLOCKED_CHECKOUT_STATUSES:
                skipped += 1
                continue
            db.session.add(Assignment(asset=a, employee=emp, assigned_by=g.user.id,
                                      handled_by=g.user.name))
            a.status = "Checked Out"
            a.updated_by = g.user.name
            log_activity("checked_out", "asset", a.id, f"{a.tag} → {emp.name} (bulk)")
            done += 1
        flash(f"Assigned {done} assets to {emp.name}"
              + (f" ({skipped} skipped)." if skipped else "."), "success")
    elif action == "transfer" and request.form.get("location_id"):
        loc_id = int(request.form["location_id"])
        for a in assets:
            db.session.add(Transfer(asset=a, from_location_id=a.location_id,
                                    to_location_id=loc_id, by_user=g.user.id,
                                    notes="Bulk transfer"))
            a.location_id = loc_id
            log_activity("transferred", "asset", a.id, f"{a.tag} (bulk)")
        flash(f"{len(assets)} assets transferred.", "success")
    elif action == "department" and request.form.get("department_id"):
        dep = db.get_or_404(Department, int(request.form["department_id"]))
        for a in assets:
            a.department_id = dep.id
            log_activity("updated", "asset", a.id, f"{a.tag} department → {dep.name}")
        flash(f"{len(assets)} assets moved to {dep.name}.", "success")
    elif action and action.startswith("condition:"):
        condition = action.split(":", 1)[1]
        if condition in ASSET_CONDITIONS:
            for a in assets:
                a.condition = condition
                log_activity("updated", "asset", a.id,
                             f"{a.tag} condition → {condition}")
            flash(f"{len(assets)} assets set to {condition}.", "success")
    db.session.commit()
    return redirect(url_for("assets.list_"))


@bp.post("/delete-all")
@perm_required("assets.manage")
def delete_all():
    """Wipe the whole register in one click, e.g. before a clean re-import.

    Page-by-page bulk delete meant doing 50 rows at a time; the button that
    posts here deletes every asset (and, via cascades and explicit deletes,
    their assignments, transfers and history) after an explicit confirm.
    """
    assets = db.session.scalars(db.select(Asset)).all()
    snaps = [_asset_snapshot(a) for a in assets]
    for a in assets:
        db.session.delete(a)
    log_activity("deleted", "asset", None, f"all assets ({len(assets)})",
                 undo_data=json.dumps({"op": "restore", "assets": snaps}))
    db.session.commit()
    _refresh_locations()
    db.session.commit()
    flash(f"All {len(assets)} assets deleted.", "success")
    return redirect(url_for("assets.list_"))


# ------------------------------------------------------- import (Excel/CSV)

IMPORT_COLS = ["tag", "name", "category / asset", "type", "serial / serial no.",
               "manufacturer", "model", "status / asset status", "condition",
               "branch", "building", "floor", "location", "room", "assigned to",
               "department / dept", "ip address", "os", "cpu", "ram", "storage",
               "hostname", "mac address", "purchase date", "warranty expiry",
               "purchase cost", "updated by", "notes"]

# Map many possible spreadsheet header names -> our canonical field.
# Keys are "normalized" (lowercased, non-alphanumerics collapsed to spaces).
HEADER_ALIASES = {
    "tag": "tag", "asset tag": "tag", "asset id": "tag", "assetid": "tag",
    "barcode qr code": "tag", "barcode": "tag", "qr code": "tag", "qr": "tag",
    "name": "name", "device name": "name", "asset name": "name", "device": "name",
    "hostname": "hostname", "host name": "hostname", "computer name": "hostname",
    "asset": "category", "category": "category", "asset category": "category",
    "type": "type", "asset type": "type", "device type": "type",
    "serial": "serial", "serial no": "serial", "serial number": "serial",
    "serialno": "serial", "serial num": "serial",
    "manufacturer": "manufacturer", "brand": "manufacturer", "make": "manufacturer",
    "model": "model",
    "status": "status", "asset status": "status", "state": "status",
    "condition": "condition",
    "branch": "branch", "site": "branch", "campus": "branch",
    "building": "building",
    "floor": "floor", "level": "floor",
    "location": "location", "location name": "location", "place": "location",
    "room": "room",
    "assigned to": "assigned_to", "assigned": "assigned_to", "assignee": "assigned_to",
    "assign to": "assigned_to", "assign": "assigned_to",
    "user": "assigned_to", "owner": "assigned_to", "holder": "assigned_to",
    "department": "department", "dept": "department",
    "ip address": "ip_address", "ip": "ip_address", "ipaddress": "ip_address",
    "os": "os", "operating system": "os", "os name": "os",
    "cpu": "cpu", "processor": "cpu",
    "ram": "ram", "memory": "ram",
    "storage": "storage", "disk": "storage", "hdd": "storage", "ssd": "storage",
    "mac": "mac_address", "mac address": "mac_address",
    "gpu": "gpu", "graphic cards": "gpu", "graphics card": "gpu",
    "graphic card": "gpu", "graphics": "gpu",
    "depreciation years": "depreciation", "depreciation": "depreciation",
    "invoice number": "invoice", "invoice": "invoice", "invoice no": "invoice",
    "purchase date": "purchase_date", "purchased": "purchase_date", "buy date": "purchase_date",
    "warranty expiry": "warranty_expiry", "warranty": "warranty_expiry",
    "warranty expiration": "warranty_expiry", "warranty end": "warranty_expiry",
    "purchase cost": "purchase_cost", "cost": "purchase_cost", "price": "purchase_cost",
    "updated by": "updated_by", "data entry": "updated_by", "entered by": "updated_by",
    "notes": "notes", "note": "notes", "remarks": "notes", "comment": "notes",
    "comments": "notes",
}


def _norm_header(h):
    out = []
    for ch in str(h or "").lower():
        out.append(ch if ch.isalnum() else " ")
    return " ".join("".join(out).split())


def _flex_date(value):
    """Parse a date written in several common formats; return date or None."""
    v = str(value or "").strip()
    if not v:
        return None
    v = v.split(" ")[0]  # drop any time portion
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y",
                "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def _xlsx_to_csv(file_storage):
    """Convert the first sheet of an .xlsx upload to a CSV string."""
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if v is None else
                    (v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else v)
                    for v in row])
    wb.close()
    return buf.getvalue()


@bp.route("/export.xlsx")
@perm_required("assets.view")
def export_xlsx():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    ws.append(EXPORT_HEADERS)
    for a in _filtered_assets(request.args):
        ws.append(_export_row(a, blank=None))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, download_name="assets.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


@bp.route("/import", methods=["GET", "POST"])
@perm_required("assets.manage")
def import_():
    preview, errors, token = None, [], None
    if request.method == "POST" and "file" in request.files:
        f = request.files["file"]
        if (f.filename or "").lower().endswith((".xlsx", ".xlsm")):
            try:
                raw = _xlsx_to_csv(f)
            except Exception:
                flash("Could not read that Excel file — is it a valid .xlsx?", "error")
                return render_template("assets/import.html", preview=None, errors=[],
                                       token=None, cols=IMPORT_COLS)
        else:
            raw = f.read().decode("utf-8-sig", errors="replace")
        token = f"import-{uuid.uuid4().hex}.csv"
        with open(os.path.join(current_app.config["UPLOAD_FOLDER"], token), "w",
                  encoding="utf-8") as out:
            out.write(raw)
        preview, errors = _validate_import(raw)
    elif request.method == "POST" and request.form.get("token"):
        token_name = os.path.basename(request.form["token"])
        path = os.path.join(current_app.config["UPLOAD_FOLDER"], token_name)
        with open(path, encoding="utf-8") as fh:
            raw = fh.read()
        rows, errors = _validate_import(raw)
        created, created_ids = _apply_asset_import(rows)
        log_activity("imported", "asset", None, f"{created} assets from file",
                     undo_data=json.dumps({"op": "delete", "asset_ids": created_ids}))
        db.session.commit()
        try:
            os.remove(path)
        except OSError:
            pass
        dups = len([r for r in rows
                    if r["_error"] and "Duplicated" in r["_error"]])
        skipped = len([r for r in rows if r["_error"]]) - dups
        _refresh_locations()
        message = f"Imported {created} assets."
        if dups:
            message += f" Duplicated Records: {dups} rows refused."
        if skipped:
            message += f" {skipped} other rows skipped."
        flash(message, "error" if dups else "success")
        return redirect(url_for("assets.list_"))
    return render_template("assets/import.html", preview=preview, errors=errors,
                           token=token, cols=IMPORT_COLS)


#: Words that mark an "Assign to" value as a PLACE (a room, lab, class,
#: office, …) rather than a person. School sheets routinely put the room a
#: shared device lives in into the Assign-to column; creating an "employee"
#: called "Copy Room" or "Grade.3.A" for those is wrong.
PLACE_WORDS = {
    "room", "rooms", "lab", "labs", "office", "offices", "class", "classes",
    "classroom", "grade", "dept", "department", "library", "clinic",
    "reception", "storage", "store", "hall", "kg",
    "صف", "غرفة", "مختبر", "مكتب", "قسم", "مكتبة", "عيادة", "استقبال",
    "مخزن", "قاعة", "جهاز",
}


#: Words that mark an "Assign to" value as a PERSON even when the rest of it
#: looks odd — titles in English and Arabic. "أستاذ إبراهيم 2" is a person
#: despite the digit; "Kg1.A" is not.
PERSON_WORDS = {
    "mr", "mrs", "ms", "miss", "dr", "eng", "coach", "sir", "madam",
    "مس", "مستر", "أستاذ", "استاذ", "دكتور", "دكتورة", "مدير", "مديرة",
    "مشرف", "مشرفة", "معلم", "معلمة", "آنسة", "انسة", "أبلة", "ابلة", "مربية",
}


def _clean_person(name):
    """Strip a leading room/desk code from a person: 'REG-03/ Mr. Saleh' ->
    'Mr. Saleh', 'Recp-01 Ms. Alaa' -> 'Ms. Alaa'. Without this the same
    person imports once per desk they sit at, each with a dirty name."""
    import re
    m = re.match(r"^\s*[A-Za-z؀-ۿ]{1,10}[-_. ]?\d+\s*[/\-–:]?\s+(.+)$",
                 name)
    return m.group(1).strip() if m and m.group(1).strip() else name


def _looks_like_place(value, row, known_places):
    """True when an "Assign to" value names a place or a class, not a person."""
    v = value.strip().lower()
    if not v:
        return False
    tokens = set(_norm_header(value).split())
    if tokens & PERSON_WORDS:
        return False                        # a title always means a person
    if v in known_places:
        return True
    for key in ("room", "location", "department", "building", "branch", "floor"):
        if v == row[key].strip().lower():
            return True
    # "المكتبة" is "مكتبة" wearing the definite article — strip a leading
    # "ال" before looking a token up.
    if any(tok in PLACE_WORDS
           or (tok.startswith("ال") and tok[2:] in PLACE_WORDS)
           for tok in tokens):
        return True
    # Class and room codes carry digits ("Kg1.A", "Grade.10.B", "10A");
    # people's names don't.
    return any(ch.isdigit() for ch in v)


def _apply_asset_import(rows):
    """Create assets from validated rows, auto-creating categories, departments
    and employees, assigning assets to their holders, and auto-generating a
    unique Asset ID whenever the file's tag is missing or duplicated."""
    import re

    from .org import next_employee_code

    # 1) auto-create categories and departments referenced in the file
    cats = {c.name.lower(): c for c in db.session.scalars(db.select(Category))}
    deps = {d.name.lower(): d for d in db.session.scalars(db.select(Department))}
    for r in rows:
        if r["_error"]:
            continue
        cn = r["category"].strip()
        if cn and cn.lower() not in cats:
            c = Category(name=cn)
            db.session.add(c)
            cats[cn.lower()] = c
        dn = r["department"].strip()
        if dn and dn.lower() not in deps:
            d = Department(name=dn)
            db.session.add(d)
            deps[dn.lower()] = d
    db.session.flush()  # assign ids to the new categories/departments

    # "Assign to" names become real employees with real assignments — not a
    # line of text in the notes. Match by name; create the ones we've never
    # seen, continuing the existing Employee ID sequence.
    emps = {e.name.strip().lower(): e
            for e in db.session.scalars(db.select(Employee))}
    known_places = {loc.name.strip().lower()
                    for loc in db.session.scalars(db.select(Location))}
    for r in rows:
        if not r["_error"]:
            for key in ("room", "location", "department", "building"):
                if r[key].strip():
                    known_places.add(r[key].strip().lower())
    seed_code = next_employee_code()
    code_match = re.match(r"^(.*?)(\d+)$", seed_code)

    def gen_emp_code():
        nonlocal code_match
        if not code_match:
            return None
        stem, num = code_match.group(1), code_match.group(2)
        code_match = re.match(r"^(.*?)(\d+)$",
                              f"{stem}{int(num) + 1:0{len(num)}d}")
        return f"{stem}{num}"

    def employee_for(name, dep):
        emp = emps.get(name.lower())
        if emp is None:
            emp = Employee(name=name, emp_code=gen_emp_code(),
                           department_id=dep.id if dep else None)
            db.session.add(emp)
            emps[name.lower()] = emp
        return emp

    # 2) seed per-prefix counters and the set of tags already in use
    used = {t.lower() for (t,) in db.session.execute(db.select(Asset.tag)).all()}
    counters = {}

    def gen_tag(cat):
        prefix = cat.tag_prefix if cat else "AST"
        n = counters.get(prefix)
        if n is None:
            n = 0
            for t in used:
                if t.startswith(prefix.lower() + "-"):
                    suf = t.rsplit("-", 1)[-1]
                    if suf.isdigit():
                        n = max(n, int(suf))
        while True:
            n += 1
            counters[prefix] = n
            cand = f"{prefix}-{n:06d}"
            if cand.lower() not in used:
                used.add(cand.lower())
                return cand

    created_assets = []
    for r in rows:
        if r["_error"]:
            continue
        cat = cats.get(r["category"].strip().lower()) if r["category"].strip() else None
        dep = deps.get(r["department"].strip().lower()) if r["department"].strip() else None
        tag = r["tag"].strip()
        if not tag or r["_auto_tag"] or tag.lower() in used:
            tag = gen_tag(cat)
        else:
            used.add(tag.lower())
        # A second room only goes into notes when the location column already
        # holds something else — when the room IS the location, noting it
        # again would just duplicate what the Location field shows.
        extra = []
        if (r["room"].strip() and r["location"].strip()
                and r["room"].strip().lower() != r["location"].strip().lower()):
            extra.append(f"Room: {r['room'].strip()}")
        # An "Assign to" that names a class or room can't become an employee,
        # but it mustn't vanish either: keep it in the notes unless it just
        # repeats one of the row's own location fields.
        # Collapse doubled spaces so "Ms.  Safa" and "Ms. Safa" are one person.
        holder = " ".join(r["assigned_to"].split())
        holder_is_place = bool(holder) and _looks_like_place(holder, r, known_places)
        if holder and not holder_is_place:
            holder = _clean_person(holder)
        if holder_is_place and holder.lower() not in {
                r[k].strip().lower()
                for k in ("room", "location", "department", "building",
                          "branch", "floor")}:
            extra.append(f"Assigned to: {holder}")
        notes = "\n".join([p for p in [r["notes"].strip(), *extra] if p]) or None
        try:
            cost = float(r["purchase_cost"]) if r["purchase_cost"].strip() else None
        except ValueError:
            cost = None
        try:
            dep_years = int(float(r["depreciation"])) if r["depreciation"].strip() else None
        except ValueError:
            dep_years = None
        a = Asset(
            tag=tag,
            name=r["name"].strip(),
            category_id=cat.id if cat else None,
            # if the sheet has no separate "type" column, reuse the category
            # (e.g. the "Asset" column: Desktop / Printer / IP Phone …)
            asset_type=(r["type"].strip() or r["category"].strip()) or None,
            serial=r["serial"].strip() or None,
            manufacturer=r["manufacturer"].strip() or None,
            model=r["model"].strip() or None,
            status=r["status"].strip() if r["status"].strip() in ASSET_STATUSES else "In Use",
            condition=r["condition"].strip() if r["condition"].strip() in ASSET_CONDITIONS else "Good",
            branch=r["branch"].strip() or None,
            building=r["building"].strip() or None,
            floor=r["floor"].strip() or None,
            location_name=(r["location"].strip() or r["room"].strip()) or None,
            updated_by=r["updated_by"].strip() or None,
            department_id=dep.id if dep else None,
            ip_address=r["ip_address"].strip() or None,
            os_name=r["os"].strip() or None,
            cpu=r["cpu"].strip() or None,
            ram=r["ram"].strip() or None,
            storage=r["storage"].strip() or None,
            hostname=(r["hostname"].strip() or r["name"].strip()) or None,
            mac_address=r["mac_address"].strip() or None,
            gpu=r["gpu"].strip() or None,
            invoice_number=r["invoice"].strip() or None,
            purchase_date=_flex_date(r["purchase_date"]),
            warranty_expiry=_flex_date(r["warranty_expiry"]),
            purchase_cost=cost,
            notes=notes,
        )
        if dep_years:
            a.depreciation_years = dep_years
        db.session.add(a)
        created_assets.append(a)
        if holder and not holder_is_place:
            db.session.add(Assignment(
                asset=a, employee=employee_for(holder, dep),
                assigned_by=g.user.id if g.get("user") else None))
    db.session.flush()      # assign ids so the import can be undone later
    return len(created_assets), [a.id for a in created_assets]


# canonical fields carried on each parsed row
_CANON = ["tag", "name", "category", "type", "serial", "manufacturer", "model",
          "status", "condition", "branch", "building", "floor", "location", "room",
          "assigned_to", "department", "ip_address", "os", "cpu", "ram", "storage",
          "hostname", "mac_address", "purchase_date", "warranty_expiry",
          "purchase_cost", "updated_by", "notes", "gpu", "depreciation",
          "invoice"]


def _validate_import(raw):
    errors = []
    reader = csv.DictReader(io.StringIO(raw))
    # map each source header to a canonical field (first match wins)
    col_of = {}
    for fn in (reader.fieldnames or []):
        canon = HEADER_ALIASES.get(_norm_header(fn))
        if canon and canon not in col_of:
            col_of[canon] = fn
    if "name" not in col_of and "tag" not in col_of and "hostname" not in col_of:
        return [], ['The file needs at least a "Name" column (a "Tag"/"Asset ID" '
                    'column is optional — IDs are generated automatically).']

    raw_rows = list(reader)
    # Asset IDs already in the register: a row reusing one is a duplicate
    # record and is refused, not silently renumbered.
    existing = {t.lower() for (t,) in db.session.execute(db.select(Asset.tag))}
    # The school's sheet leaves the Asset ID column empty, so importing the
    # same file twice used to double the register: every row got a fresh
    # generated ID. A row without an ID whose name+serial are already in the
    # register is the same device coming back — refuse it too.
    existing_pairs = {
        ((n or "").strip().lower(), (s or "").strip().lower())
        for n, s in db.session.execute(db.select(Asset.name, Asset.serial))}

    seen, rows = set(), []
    for i, rec in enumerate(raw_rows, start=2):
        row = {c: _row_val(rec, col_of, c) for c in _CANON}
        # a usable name: fall back to hostname, then tag, then category+line
        if not row["name"]:
            row["name"] = (row["hostname"] or row["tag"]
                           or (f"{row['category']} {i - 1}".strip() if row["category"]
                               else ""))
        err = None
        if not row["name"]:
            err = "row is empty (no name)"
        # Asset ID rules: a new ID imports; a known or repeated ID is refused;
        # a missing ID is generated by the system (last ID + 1).
        tag = row["tag"].lower()
        if not err and tag:
            if tag in existing:
                err = (f'Duplicated Records — Asset ID "{row["tag"]}" '
                       "already exists in the register")
            elif tag in seen:
                err = (f'Duplicated Records — Asset ID "{row["tag"]}" '
                       "appears more than once in the file")
            else:
                seen.add(tag)
        elif (not err and row["serial"].strip()
                and (row["name"].lower(),
                     row["serial"].lower()) in existing_pairs):
            # Only the Asset ID is a key; names may repeat freely. This
            # extra check needs a real serial to fire — the same name WITH
            # the same serial is the same physical device coming back.
            err = (f'Duplicated Records — "{row["name"]}" with this serial is '
                   "already in the register (give the row a new Asset ID if "
                   "it really is another device)")
        row["_auto_tag"] = not tag
        row["_line"], row["_error"] = i, err
        if err:
            errors.append(f"Row {i}: {err}")
        rows.append(row)
    return rows, errors


def _row_val(rec, col_of, canon):
    src = col_of.get(canon)
    if not src:
        return ""
    return (rec.get(src, "") or "").strip()
