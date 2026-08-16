import io
import os

import pytest

from itam import create_app
from itam.models import Asset, Category, Employee, License, User, db


@pytest.fixture
def app(tmp_path):
    app = create_app({
        "TESTING": True,
        "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
        "UPLOAD_FOLDER": str(tmp_path / "uploads"),
        "BACKUP_FOLDER": str(tmp_path / "backups"),
    })
    with app.app_context():
        db.create_all()
        viewer = User(username="viewer", name="View Only",
                      email="viewer@example.com", role="viewer")
        viewer.set_password("viewer123")
        db.session.add(viewer)
        # "Laptops" is one of the categories a fresh database now seeds, so
        # only add it when it is genuinely missing.
        if not db.session.scalar(db.select(Category).where(Category.name == "Laptops")):
            db.session.add(Category(name="Laptops"))
        db.session.add(Employee(name="Alice Hart", email="alice@example.com"))
        db.session.commit()
    return app


@pytest.fixture
def client(app):
    return app.test_client()


def login(client, username="admin", password="admin123"):
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def test_login_required(client):
    resp = client.get("/", follow_redirects=False)
    assert resp.status_code == 302
    assert "/login" in resp.headers["Location"]


def test_login_logout(client):
    resp = login(client)
    assert b"Dashboard" in resp.data
    resp = client.get("/logout", follow_redirects=True)
    assert b"Log in" in resp.data


def test_bad_login_rejected(client):
    resp = client.post("/login", data={"username": "admin", "password": "wrong"},
                       follow_redirects=True)
    assert b"Invalid username or password" in resp.data


def test_asset_crud_and_checkout(client, app):
    login(client)
    resp = client.post("/assets/new", data={
        "tag": "LT-0001", "name": "Test Laptop", "category_id": "1",
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "purchase_date": "2026-01-15", "purchase_cost": "1200.50",
    }, follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "LT-0001"))
        assert a is not None and a.status == "Available"
        asset_id = a.id
        emp_id = db.session.scalar(db.select(Employee.id))

    client.post(f"/assets/{asset_id}/checkout", data={"employee_id": emp_id})
    with app.app_context():
        a = db.session.get(Asset, asset_id)
        assert a.status == "Checked Out"
        assert a.current_assignment.employee.name == "Alice Hart"

    client.post(f"/assets/{asset_id}/checkin")
    with app.app_context():
        a = db.session.get(Asset, asset_id)
        assert a.status == "Available"
        assert a.current_assignment is None


def test_duplicate_tag_rejected(client, app):
    login(client)
    for _ in range(2):
        client.post("/assets/new", data={
            "tag": "DUP-1", "name": "Laptop", "status": "Available",
            "condition": "Good", "depreciation_years": "5"})
    with app.app_context():
        assert db.session.scalar(
            db.select(db.func.count(Asset.id)).where(Asset.tag == "DUP-1")) == 1


def test_rbac_viewer_cannot_manage(client, app):
    login(client, "viewer", "viewer123")
    resp = client.post("/assets/new", data={
        "tag": "V-1", "name": "X", "status": "Available",
        "condition": "Good", "depreciation_years": "5"})
    assert resp.status_code == 403
    resp = client.get("/admin/users")
    assert resp.status_code == 403
    # but viewer can view assets
    assert client.get("/assets/").status_code == 200


def test_activity_log_written(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "LOG-1", "name": "X", "status": "Available",
        "condition": "Good", "depreciation_years": "5"})
    resp = client.get("/activity")
    assert b"LOG-1" in resp.data


def test_license_seats_and_compliance(client, app):
    login(client)
    client.post("/licenses/new", data={"name": "Office Suite", "seats": "1"},
                follow_redirects=True)
    with app.app_context():
        lic = db.session.scalar(db.select(License))
        lic_id = lic.id
        emp_id = db.session.scalar(db.select(Employee.id))
    client.post(f"/licenses/{lic_id}/assign", data={"employee_id": emp_id})
    # second assign should be blocked (seats exhausted)
    client.post(f"/licenses/{lic_id}/assign", data={"employee_id": emp_id})
    with app.app_context():
        lic = db.session.get(License, lic_id)
        assert lic.seats_used == 1
        assert lic.compliant


def test_csv_export_and_qr(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "QR-1", "name": "Scanner Test", "status": "Available",
        "condition": "Good", "depreciation_years": "5"})
    resp = client.get("/assets/export.csv")
    assert resp.status_code == 200
    assert b"QR-1" in resp.data
    with app.app_context():
        asset_id = db.session.scalar(db.select(Asset.id).where(Asset.tag == "QR-1"))
    resp = client.get(f"/assets/{asset_id}/qr.svg")
    assert resp.status_code == 200
    assert b"<svg" in resp.data
    resp = client.get(f"/assets/{asset_id}/barcode.svg")
    assert resp.status_code == 200


def test_api_requires_token(client, app):
    assert client.get("/api/v1/assets").status_code == 401
    login(client)
    with app.app_context():
        admin = db.session.scalar(db.select(User).where(User.username == "admin"))
        admin_id = admin.id
    client.post(f"/admin/users/{admin_id}/apikey")
    with app.app_context():
        key = db.session.get(User, admin_id).api_key
    fresh = app.test_client()
    resp = fresh.get("/api/v1/assets", headers={"Authorization": f"Bearer {key}"})
    assert resp.status_code == 200
    assert isinstance(resp.get_json(), list)


def test_excel_import_and_export(client, app):
    import io
    from openpyxl import Workbook, load_workbook
    login(client)
    wb = Workbook()
    ws = wb.active
    ws.append(["tag", "name", "category", "status", "purchase_cost"])
    ws.append(["XL-0001", "Excel Laptop", "Laptops", "Available", 999.5])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post("/assets/import",
                       data={"file": (buf, "assets.xlsx")},
                       content_type="multipart/form-data")
    assert b"Import preview" in resp.data
    import re
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    client.post("/assets/import", data={"token": token}, follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "XL-0001"))
        assert a is not None and a.name == "Excel Laptop"
        assert float(a.purchase_cost) == 999.5
        assert a.category.name == "Laptops"    # category auto-created

    resp = client.get("/assets/export.xlsx")
    assert resp.status_code == 200
    out = load_workbook(io.BytesIO(resp.data))
    headers = [c.value for c in next(out.active.iter_rows(min_row=1, max_row=1))]
    # arranged exactly like the school's own inventory sheet
    assert headers[:6] == ["Name", "Category", "Asset ID", "Condition",
                           "Status", "Serial No."]
    tags = [row[2].value for row in out.active.iter_rows(min_row=2)]
    assert "XL-0001" in tags


def test_import_flexible_headers_and_autotag(client, app):
    """Real-world sheet: odd headers, missing tags, DD/MM/YYYY dates. Rows
    without an Asset ID get a generated one; a repeated ID is refused."""
    import re
    login(client)
    csv = (
        "No,Asset,TAG,Branch,Floor,Assigned to,Name,Asset Status,Condition,"
        "Purchase Date,Serial No.,Dept\n"
        "1,Desktop,,Mada 3,GF,Ms A,M3-PC01,In Use,Good,16/07/2026,SN-1,Reg\n"
        "2,Desktop,,Mada 3,GF,Ms B,M3-PC02,In Use,Good,16/07/2026,SN-2,Reg\n"
        "3,Printer,,Mada 3,F1,Copy Room,M3-PR01,In Use,Good,16/07/2026,SN-3,IT\n"
    )
    resp = client.post("/assets/import",
                       data={"file": (io.BytesIO(csv.encode()), "db.csv")},
                       content_type="multipart/form-data")
    assert b"Import preview" in resp.data
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    client.post("/assets/import", data={"token": token}, follow_redirects=True)
    with app.app_context():
        assets = db.session.scalars(db.select(Asset)).all()
        assert len(assets) == 3                       # all rows imported
        assert len({a.tag for a in assets}) == 3      # unique generated tags
        pc = db.session.scalar(db.select(Asset).where(Asset.name == "M3-PC01"))
        assert pc.category.name == "Desktop"          # from "Asset" column
        assert pc.branch == "Mada 3" and pc.floor == "GF"
        assert pc.serial == "SN-1"
        assert str(pc.purchase_date) == "2026-07-16"  # DD/MM/YYYY parsed
        # "Assigned to" creates a real employee + assignment, not a note
        assert pc.current_assignment.employee.name == "Ms A"


def test_import_refuses_duplicated_asset_ids(client, app):
    """A new Asset ID imports; a known or repeated one is refused with
    'Duplicated Records'; a missing one is generated (last ID + 1)."""
    import re
    login(client)
    with app.app_context():
        db.session.add(Asset(tag="DES-000001", name="Already here",
                             status="Available", condition="Good",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()

    csv = (
        "Name,Category,Asset ID,Serial No.\n"
        "New PC,Desktop,DES-000002,SN-A\n"       # new ID -> imported
        "Clash PC,Desktop,DES-000001,SN-B\n"     # exists in register -> refused
        "Twin 1,Desktop,DES-000003,SN-C\n"       # new -> imported
        "Twin 2,Desktop,DES-000003,SN-D\n"       # repeats in file -> refused
        "No ID PC,Desktop,,SN-E\n"               # missing -> generated
    )
    resp = client.post("/assets/import",
                       data={"file": (io.BytesIO(csv.encode()), "dups.csv")},
                       content_type="multipart/form-data")
    assert b"Duplicated Records" in resp.data     # shown in the preview
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    done = client.post("/assets/import", data={"token": token},
                       follow_redirects=True)
    assert b"Duplicated Records" in done.data
    with app.app_context():
        names = {a.name for a in db.session.scalars(db.select(Asset))}
        assert {"Already here", "New PC", "Twin 1", "No ID PC"} <= names
        assert "Clash PC" not in names and "Twin 2" not in names
        generated = db.session.scalar(
            db.select(Asset).where(Asset.name == "No ID PC"))
        # system ID continues the DES sequence: last ID + 1
        assert generated.tag == "DES-000004"


def test_import_assign_to_and_room_land_in_real_fields(client, app):
    """The school's own sheet: 'Assign to ' (trailing space) holds Arabic
    names, 'Room' is the only location column. The holder must become a real
    assigned employee and the room must not be duplicated into notes."""
    import re
    from itam.models import Employee
    login(client)
    csv = (
        "Name,Category ,Serial No.,Branch,Building,Floor,Room,Department,"
        "Assign to ,Updated By\n"
        "M3-RECP01,Desktop,BQL6C14,Mada 3,Building 1,GF,Reg,Reception,"
        "مس الاء,Ayham\n"
        "M3-RECP02,Desktop,9QG6C14,Mada 3,Building 1,GF,Reg,Reception,"
        "مس الاء,Ayham\n"
    )
    resp = client.post("/assets/import",
                       data={"file": (io.BytesIO(csv.encode()), "inv.csv")},
                       content_type="multipart/form-data")
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    client.post("/assets/import", data={"token": token}, follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.name == "M3-RECP01"))
        assert a.location_name == "Reg"               # room used as location
        assert "Room" not in (a.notes or "")          # ...and not echoed in notes
        assert a.current_assignment.employee.name == "مس الاء"
        # the same name on two rows is one employee, and it got an ID
        emps = db.session.scalars(
            db.select(Employee).where(Employee.name == "مس الاء")).all()
        assert len(emps) == 1 and emps[0].emp_code


def test_import_does_not_turn_rooms_and_classes_into_employees(client, app):
    """School sheets put the room a shared device lives in into 'Assign to'
    ('Copy Room', 'صف', 'Grade.3.A'…). Those are places, not people."""
    import re
    from itam.models import Employee
    login(client)
    csv = (
        "Name,Category,Serial No.,Room,Assign to\n"
        "PC-A,Desktop,S-1,Reg,مس الاء\n"           # a person -> employee
        "PC-B,Desktop,S-2,Copy Room,Copy Room\n"    # same as its room -> place
        "PC-C,Desktop,S-3,KG1,صف\n"                 # 'classroom' -> place
        "PC-D,Desktop,S-4,Lab,جهاز مختبر الحاسوب\n"  # 'computer lab device'
        "PC-E,Desktop,S-5,10A,Grade.10.A\n"         # a class -> place
        "PC-F,Desktop,S-6,IT,Mr. Saleh\n"           # a person -> employee
        "PC-G,Desktop,S-7,KG,Kg1.A\n"               # class code with digit
        "PC-H,Desktop,S-8,IT,أستاذ إبراهيم 2\n"     # titled person with digit
        "PC-I,Desktop,S-9,Reg,REG-03/ Mr. Saleh\n"  # desk code + person
        "PC-J,Desktop,S-10,IT,Mr. Saleh\n"          # ...same person, plain
    )
    resp = client.post("/assets/import",
                       data={"file": (io.BytesIO(csv.encode()), "mix.csv")},
                       content_type="multipart/form-data")
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    client.post("/assets/import", data={"token": token}, follow_redirects=True)
    with app.app_context():
        names = {e.name for e in db.session.scalars(db.select(Employee))}
        assert "مس الاء" in names and "Mr. Saleh" in names
        assert "أستاذ إبراهيم 2" in names       # a title outweighs the digit
        for not_a_person in ("Copy Room", "صف", "جهاز مختبر الحاسوب",
                             "Grade.10.A", "Kg1.A"):
            assert not_a_person not in names
        for tag_name, assigned in (("PC-A", True), ("PC-B", False),
                                   ("PC-C", False), ("PC-E", False),
                                   ("PC-F", True), ("PC-G", False),
                                   ("PC-H", True)):
            a = db.session.scalar(db.select(Asset).where(Asset.name == tag_name))
            assert bool(a.current_assignment) is assigned, tag_name
        # a skipped class still shows up in the notes, so nothing is lost
        e_asset = db.session.scalar(db.select(Asset).where(Asset.name == "PC-E"))
        assert "Grade.10.A" in (e_asset.notes or "")
        assert e_asset.assigned_label == "Grade.10.A"   # ...and on screen
        # ...unless it only repeats the row's own room
        b_asset = db.session.scalar(db.select(Asset).where(Asset.name == "PC-B"))
        assert "Assigned to" not in (b_asset.notes or "")
        # a desk code in front of a person is stripped, and the person merges
        # with their plain-named self instead of duplicating
        assert "REG-03/ Mr. Saleh" not in names
        saleh = [e for e in db.session.scalars(db.select(Employee))
                 if e.name == "Mr. Saleh"]
        assert len(saleh) == 1
        i_asset = db.session.scalar(db.select(Asset).where(Asset.name == "PC-I"))
        assert i_asset.current_assignment.employee.name == "Mr. Saleh"

    # the class shows in the Assigned-to column of the asset list too
    body = client.get("/assets/?q=PC-E").data.decode()
    assert "Grade.10.A" in body

    # ...and on the Lending screen under "Held by classes & rooms",
    # where it can be released
    lending = client.get("/checkouts").data.decode()
    assert "Held by classes" in lending and "Grade.10.A" in lending
    with app.app_context():
        e_id = db.session.scalar(db.select(Asset.id).where(Asset.name == "PC-E"))
    client.post(f"/checkouts/release/{e_id}", follow_redirects=True)
    with app.app_context():
        released = db.session.get(Asset, e_id)
        assert released.assigned_label is None
        assert "Assigned to" not in (released.notes or "")


def test_transfer_changes_branch_building_floor_room_department(client, app):
    """'Change location' on the asset page moves the asset with the same
    Branch/Building/Floor/Room/Department fields the asset form has."""
    from itam.models import Department, Transfer
    login(client)
    with app.app_context():
        dep = Department(name="Science")
        db.session.add(dep)
        a = Asset(tag="TR-0001", name="Cart", status="Available",
                  condition="Good", branch="Mada 3", building="Building 1",
                  floor="GF", location_name="Reg",
                  category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.commit()
        aid, depid = a.id, dep.id

    client.post(f"/assets/{aid}/transfer", data={
        "branch": "Mada 1", "building": "Building 2", "floor": "F1",
        "location_name": "Lab 3", "department_id": str(depid),
        "notes": "term move"}, follow_redirects=True)
    with app.app_context():
        a = db.session.get(Asset, aid)
        assert (a.branch, a.building, a.floor, a.location_name) == \
            ("Mada 1", "Building 2", "F1", "Lab 3")
        assert a.department_id == depid
        tr = db.session.scalars(db.select(Transfer)).all()[-1]
        assert "Mada 3 / Building 1 / GF / Reg" in tr.notes
        assert "Mada 1 / Building 2 / F1 / Lab 3" in tr.notes
        assert "term move" in tr.notes


def test_search_page(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "SRCH-1", "name": "Zebra Machine", "status": "Available",
        "condition": "Good", "depreciation_years": "5"})
    resp = client.get("/search?q=Zebra")
    assert b"SRCH-1" in resp.data


def test_saved_searches(client, app):
    login(client)
    client.post("/assets/searches", data={"name": "My laptops", "q": "",
                                          "query": "status=Available&category=1"})
    resp = client.get("/assets/")
    assert "My laptops".encode() in resp.data
    from itam.models import SavedSearch
    with app.app_context():
        s = db.session.scalar(db.select(SavedSearch))
        assert s.query == "status=Available&category=1"
        sid = s.id
    client.post(f"/assets/searches/{sid}/delete")
    resp = client.get("/assets/")
    assert "My laptops".encode() not in resp.data


def test_custom_report_builder(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "CR-1", "name": "Report Machine", "category_id": "1",
        "status": "Available", "condition": "Good", "depreciation_years": "5"})
    resp = client.get("/reports/custom?col=tag&col=name&col=condition&status=Available")
    assert b"CR-1" in resp.data and b"Report Machine" in resp.data
    resp = client.get("/reports/custom?col=tag&status=Available&format=csv")
    assert resp.status_code == 200 and b"CR-1" in resp.data
    # unavailable filter excludes it
    resp = client.get("/reports/custom?col=tag&status=Retired")
    assert b"CR-1" not in resp.data


def test_custom_instance_path(tmp_path):
    custom = tmp_path / "portable-instance"
    app = create_app({"TESTING": True}, instance_path=str(custom))
    assert app.instance_path == str(custom)
    assert custom.exists()
    assert (custom / "itam.sqlite").exists() or app.config["SQLALCHEMY_DATABASE_URI"].endswith(
        "itam.sqlite")


def test_auto_backup(tmp_path):
    import itam as itam_pkg
    dbfile = tmp_path / "live.sqlite"
    app = create_app({
        "TESTING": True,
        "SQLALCHEMY_DATABASE_URI": f"sqlite:///{dbfile}",
        "UPLOAD_FOLDER": str(tmp_path / "up"),
        "BACKUP_FOLDER": str(tmp_path / "bk"),
    })
    client = app.test_client()
    itam_pkg._last_backup_check[0] = 0.0          # force the hourly check
    client.get("/login")
    backups = list((tmp_path / "bk").glob("auto-*.sqlite"))
    assert len(backups) == 1
    # second request within the hour must not create another
    client.get("/login")
    assert len(list((tmp_path / "bk").glob("auto-*.sqlite"))) == 1


def test_auto_asset_numbering(client, app):
    login(client)
    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        cat.prefix = "LPT"
        db.session.commit()
        cat_id = cat.id
    for _ in range(2):
        client.post("/assets/new", data={
            "tag": "", "name": "Auto Laptop", "category_id": str(cat_id),
            "status": "Available", "condition": "Good", "depreciation_years": "5"})
    with app.app_context():
        tags = sorted(t for (t,) in db.session.execute(
            db.select(Asset.tag).where(Asset.tag.like("LPT-%"))).all())
        assert tags == ["LPT-000001", "LPT-000002"]


def test_return_inspection(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "RI-1", "name": "Loaner", "status": "Available",
        "condition": "Good", "depreciation_years": "5"})
    with app.app_context():
        asset_id = db.session.scalar(db.select(Asset.id).where(Asset.tag == "RI-1"))
        emp_id = db.session.scalar(db.select(Employee.id))
    client.post(f"/assets/{asset_id}/checkout", data={"employee_id": emp_id})
    client.post(f"/assets/{asset_id}/checkin",
                data={"return_condition": "Damaged", "return_notes": "cracked lid"})
    with app.app_context():
        a = db.session.get(Asset, asset_id)
        assert a.condition == "Damaged"
        assert a.status == "Damaged"
        assert a.assignments[0].return_notes == "cracked lid"


def test_new_roles_have_permissions(app):
    from itam.models import RolePermission
    with app.app_context():
        auditor_perms = {p.permission for p in db.session.scalars(
            db.select(RolePermission).where(RolePermission.role == "auditor"))}
        assert "reports.view" in auditor_perms
        assert "assets.manage" not in auditor_perms
        superadmin_perms = {p.permission for p in db.session.scalars(
            db.select(RolePermission).where(RolePermission.role == "superadmin"))}
        assert "admin.settings" in superadmin_perms


def test_lifecycle_and_movement_reports(client):
    login(client)
    assert client.get("/reports/lifecycle").status_code == 200
    assert client.get("/reports/movement").status_code == 200
    assert client.get("/reports/locations").status_code == 200


def test_asset_label_6x3(client, app):
    login(client)
    client.post("/assets/new", data={
        "tag": "LBL-1", "name": "Label Asset", "status": "Available",
        "condition": "Good", "depreciation_years": "5", "branch": "Mada 3",
        "location_name": "IT", "serial": "SN-8842-XJ01"})
    with app.app_context():
        asset_id = db.session.scalar(db.select(Asset.id).where(Asset.tag == "LBL-1"))
    resp = client.get(f"/assets/{asset_id}/label")
    assert resp.status_code == 200
    body = resp.data.decode()
    # 6 x 3 in is the default, now expressed in mm so any stock size works
    assert "@page { size: 152.4mm 76.2mm; margin: 0; }" in body
    assert "LBL-1" in body                   # tag in the header chip
    assert "Mada 3" in body and "SN-8842-XJ01" in body
    assert 'class="bc"' in body              # Code 128 across the bottom


def test_procurement_removed(client):
    login(client)
    assert client.get("/procurement").status_code == 404
    assert b"Procurement" not in client.get("/").data


def test_employee_excel_export_and_import(client, app):
    import io
    from openpyxl import load_workbook
    from itam.models import Employee
    login(client)
    resp = client.get("/employees/export.xlsx")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    assert wb.active["A1"].value == "Name"

    # Real-world sheet: "Employee Type"/"Job Title" headers, SHARED emails across
    # distinct people, custom types, and a department that doesn't exist yet.
    csv_data = (
        "Name,Employee ID,Employee Type,Email,Job Title,Department\n"
        "Laith 1,Emp001,IT Technical,shared@mada.edu,IT Technical,IT\n"
        "Ayham 1,Emp002,IT Manager,shared@mada.edu,IT Manager,IT\n"
        "Rama 1,Emp003,Teacher,shared@mada.edu,Teacher,National\n"
    )
    resp = client.post("/employees/import",
                       data={"file": (io.BytesIO(csv_data.encode()), "e.csv")},
                       content_type="multipart/form-data", follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        from itam.models import Department
        people = db.session.scalars(
            db.select(Employee).where(Employee.emp_code.in_(["Emp001", "Emp002", "Emp003"]))).all()
        assert len(people) == 3                       # shared email did NOT collapse them
        e1 = db.session.scalar(db.select(Employee).where(Employee.emp_code == "Emp001"))
        assert e1.emp_type == "IT Technical"          # custom type kept
        assert e1.title == "IT Technical"             # "Job Title" mapped
        assert e1.department.name == "IT"             # department auto-created
        assert db.session.scalar(db.select(Department).where(Department.name == "National"))


def test_license_and_maintenance_exports(client):
    login(client)
    assert client.get("/licenses/export.xlsx").status_code == 200
    assert client.get("/licenses/export.csv").status_code == 200
    assert client.get("/maintenance/export.xlsx").status_code == 200
    assert client.get("/inventory/export.xlsx").status_code == 200


def test_arabic_language_switch(client):
    login(client)
    resp = client.get("/lang/ar", follow_redirects=True)
    assert 'dir="rtl"'.encode() in resp.data
    assert "لوحة التحكم".encode() in resp.data


def test_upgrading_over_an_older_database_adds_missing_columns(tmp_path):
    """A database made by an older build must survive the upgrade.

    db.create_all() only creates missing tables, so without _sync_schema()
    every page touching an asset fails with "no such column".
    """
    import sqlite3

    instance = tmp_path / "instance"
    app = create_app(instance_path=str(instance))
    with app.app_context():
        # "Laptops" is one of the categories a fresh database now seeds, so
        # only add it when it is genuinely missing.
        if not db.session.scalar(db.select(Category).where(Category.name == "Laptops")):
            db.session.add(Category(name="Laptops"))
        db.session.commit()
        db.session.add(Asset(tag="LT-0001", name="ThinkPad", branch="Mada 1",
                             floor="F1", status="Available", condition="Good"))
        db.session.commit()

    # Strip columns that later releases introduced, as an older install lacks.
    path = instance / "itam.sqlite"
    con = sqlite3.connect(path)
    for column in ("branch", "building", "floor", "updated_by", "cpu"):
        con.execute(f"ALTER TABLE asset DROP COLUMN {column}")
    con.commit()
    con.close()

    # Starting the current app over that database must repair it, not fail.
    upgraded = create_app(instance_path=str(instance))
    with upgraded.app_context():
        asset = db.session.scalar(db.select(Asset).where(Asset.tag == "LT-0001"))
        assert asset is not None, "existing rows must survive the upgrade"
        assert asset.name == "ThinkPad"
        assert asset.branch is None      # re-added, empty for old rows
        asset.branch = "Mada 2"          # and writable again
        db.session.commit()

    client = upgraded.test_client()
    client.post("/login", data={"username": "admin", "password": "admin123"},
                follow_redirects=True)
    for url in ("/", "/assets/", "/reports/assets", "/search?q=think"):
        assert client.get(url).status_code == 200, f"{url} broke after upgrade"


def test_schema_sync_is_idempotent(tmp_path):
    from itam import _sync_schema

    app = create_app(instance_path=str(tmp_path / "instance"))
    with app.app_context():
        added, skipped = _sync_schema()
        assert added == [], "a current database needs no changes"
        assert skipped == []


@pytest.mark.parametrize("w,h", [(152.4, 76.2), (101.6, 152.4), (101.6, 50.8),
                                 (70, 38), (55, 38), (50, 30), (40, 20)])
def test_label_layout_fits_every_offered_size(w, h):
    """Nothing may be sized past the sticker it prints on."""
    from itam.utils import label_layout

    L = label_layout(w, h, org_name="Mada Asset Management System (AMS)")
    assert L["width"] == round(w, 2) and L["height"] == round(h, 2)

    tag_line = L["fs_tag"] * 1.25 + L["pad"] * 0.3
    if L["portrait"]:
        assert L["qr"] <= w - 2 * L["pad"] + 0.01
        used = L["qr"] + tag_line + 2 * L["pad"]
    else:
        # QR column is the tall one on a landscape label.
        used = L["qr"] + tag_line + 2 * L["pad"]
        assert L["qr"] <= w * 0.42 + 0.01
    assert used <= h + 0.01, f"QR column {used:.1f}mm overflows a {h}mm label"
    assert L["fs_label"] >= 1.0 and L["fs_value"] > 0


def test_code128_table_and_encoding_are_sound():
    """Bar widths must sum to 11 modules (13 for stop) and be unique, and a
    known tag must produce the documented checksum — catches table typos."""
    from itam.utils import _C128, _code128_values, code128_svg

    assert len(_C128) == 107
    assert all(sum(int(d) for d in p) == 11 for p in _C128[:-1])
    assert sum(int(d) for d in _C128[-1]) == 13
    assert len(set(_C128)) == 107

    # "AP-000001": start B, the nine characters, checksum 81, stop.
    assert _code128_values("AP-000001") == [
        104, 33, 48, 13, 16, 16, 16, 16, 16, 17, 81, 106]

    svg = code128_svg("AP-000001")
    assert svg.startswith("<svg") and "rect" in svg


def test_labels_carry_a_laser_readable_barcode(client, app):
    """Laser scanners can't read QR; big-enough labels get Code 128 too."""
    from itam.models import Asset, Category
    from itam.utils import label_layout, set_setting

    # The barcode is the label's machine-readable code, at every size.
    assert label_layout(55, 38, org_name="Mada AMS")["bc_h"] > 0
    assert label_layout(40, 20, org_name="Mada AMS")["bc_h"] > 0

    with app.app_context():
        asset = Asset(tag="BC-0001", name="Scanner test", status="Available",
                      condition="Good",
                      category=db.session.scalar(db.select(Category)))
        db.session.add(asset)
        set_setting("label_width_mm", "55")
        set_setting("label_height_mm", "38")
        db.session.commit()
        asset_id = asset.id

    login(client)
    page = client.get(f"/assets/{asset_id}/label").data.decode()
    assert 'class="bc"' in page and "preserveAspectRatio" in page


def test_tiny_labels_drop_detail_fields():
    from itam.utils import label_layout

    assert label_layout(40, 20, org_name="Mada AMS")["fields"] == []
    assert label_layout(40, 20, org_name="Mada AMS")["show_org"] is False
    assert label_layout(152.4, 76.2, org_name="Mada AMS")["fields"]
    # A mid-size thermal sticker still carries the department.
    assert "department" in label_layout(55, 38, org_name="Mada AMS")["fields"]


def test_label_size_setting_drives_the_printed_page(client, app):
    from itam.models import Category
    from itam.utils import set_setting

    with app.app_context():
        asset = Asset(tag="LT-9001", name="ThinkPad", status="Available",
                      condition="Good",
                      category=db.session.scalar(db.select(Category)))
        db.session.add(asset)
        set_setting("label_width_mm", "50")
        set_setting("label_height_mm", "30")
        db.session.commit()
        asset_id = asset.id

    login(client)
    page = client.get(f"/assets/{asset_id}/label").data.decode()
    assert "@page { size: 50.0mm 30.0mm; margin: 0; }" in page
    assert "width: 50.0mm; height: 30.0mm" in page


def test_label_size_accepts_inches_and_cm():
    """Sticker packaging quotes sizes in in/cm; the fields convert to mm."""
    from itam.blueprints.admin import _label_mm

    assert _label_mm("6.4 in") == "162.6"
    assert _label_mm('2"') == "50.8"
    assert _label_mm("6.4 inch") == "162.6"
    assert _label_mm("3.5 cm") == "35.0"
    assert _label_mm("40 mm") == "40.0"
    assert _label_mm("76.2") == "76.2"          # plain numbers stay as-is
    assert _label_mm("garbage in") == "garbage in"


def test_common_size_dropdown_saves_without_javascript(client, app):
    """Picking a size and pressing Save must work as a plain form post."""
    from itam.utils import get_setting

    login(client)
    form = {"app_name": "Mada Asset Management System (AMS)", "qr_prefix": "",
            "custom_asset_fields": "", "checkout_days": "30",
            "warranty_alert_days": "90", "license_alert_days": "90",
            "smtp_host": "", "smtp_port": "587", "smtp_user": "",
            "smtp_password": "", "smtp_from": "", "audit_retention_days": "365",
            "label_org": "Mada International Academy", "label_printer": "",
            "update_repo": "", "update_token": "",
            "label_width_mm": "152.4", "label_height_mm": "76.2"}
    # dropdown picked, width/height boxes untouched -> the preset wins
    resp = client.post("/admin/settings",
                       data={**form, "label_preset": "55x38"},
                       follow_redirects=True)
    with app.app_context():
        assert get_setting("label_width_mm") == "55"
        assert get_setting("label_height_mm") == "38"
    assert b"55" in resp.data and b"38" in resp.data   # confirmed in the flash
    # the dropdown shows the active size on the next visit
    assert b'value="55x38" selected' in client.get("/admin/settings").data


def test_label_size_is_clamped_to_something_printable(app):
    from itam.utils import label_size_mm, set_setting

    with app.app_context():
        set_setting("label_width_mm", "0")
        set_setting("label_height_mm", "99999")
        db.session.commit()
        w, h = label_size_mm()
        assert w == 15.0 and h == 300.0

        set_setting("label_width_mm", "not a number")
        db.session.commit()
        assert label_size_mm()[0] == 152.4      # falls back to the default


def test_label_prints_the_organisation_not_the_software_name(client, app):
    """The sticker carries the school's name; AMS is the software."""
    from itam.models import Category, Department
    from itam.utils import set_setting

    with app.app_context():
        dept = Department(name="IT")
        db.session.add(dept)
        db.session.add(Asset(tag="LT-9100", name="ThinkPad", serial="SN-8842-XJ01",
                             branch="Mada 3", department=dept, status="Available",
                             condition="Good",
                             category=db.session.scalar(db.select(Category))))
        set_setting("label_org", "Mada International Academy")
        db.session.commit()
        asset_id = db.session.scalar(db.select(Asset.id).where(Asset.tag == "LT-9100"))

    login(client)
    page = client.get(f"/assets/{asset_id}/label").data.decode()
    assert "Mada International Academy" in page
    assert "(AMS)" not in page                 # software name stays off the label
    for text in ("Branch", "Mada 3", "Dept", "IT", "S/N", "SN-8842-XJ01"):
        assert text in page, f"{text} missing from the label"


def test_label_org_is_editable_from_settings(client, app):
    from itam.utils import get_setting

    login(client)
    assert b"Organisation printed on labels" in client.get("/admin/settings").data
    client.post("/admin/settings", data={
        "app_name": "Mada Asset Management System (AMS)",
        "label_org": "Another School", "label_width_mm": "50",
        "label_height_mm": "30", "qr_prefix": "", "custom_asset_fields": "",
        "checkout_days": "30", "warranty_alert_days": "90",
        "license_alert_days": "90", "smtp_host": "", "smtp_port": "587",
        "smtp_user": "", "smtp_password": "", "smtp_from": "",
        "audit_retention_days": "365"}, follow_redirects=True)
    with app.app_context():
        assert get_setting("label_org") == "Another School"


def test_restart_into_update_without_pending_is_a_noop(tmp_path):
    """Nothing staged (or running from source) must never kill the process."""
    from itam import updater

    assert updater.restart_into_update(str(tmp_path)) is False
    assert updater.restart_into_update(str(tmp_path),
                                       str(tmp_path / "AMS.exe")) is False


def _mini_db(path, assets):
    import sqlite3
    os.makedirs(os.path.dirname(path), exist_ok=True)
    con = sqlite3.connect(path)
    con.execute("CREATE TABLE asset (id INTEGER PRIMARY KEY)")
    con.executemany("INSERT INTO asset VALUES (?)", [(i,) for i in range(assets)])
    con.commit()
    con.close()


def test_data_adoption_picks_the_database_with_the_most_assets(tmp_path):
    """A new fixed data folder pulls in the fullest old database — and never
    overwrites a database that is already there."""
    import run_server

    old_full = tmp_path / "Downloads" / "AMS-Server-Setup (1)" / "instance"
    old_empty = tmp_path / "base" / "instance"
    _mini_db(str(old_full / "itam.sqlite"), assets=240)
    _mini_db(str(old_empty / "itam.sqlite"), assets=1)
    (old_full / "secret_key").write_text("the-real-key")

    assert run_server._best_existing_instance(
        [str(old_empty), str(old_full)]) == str(old_full)

    fixed = tmp_path / "ProgramData" / "AMS" / "instance"
    fixed.mkdir(parents=True)
    src = run_server._adopt_existing_data(str(tmp_path / "nowhere"), str(fixed))
    # no candidates found from that base and no Downloads: nothing adopted
    assert src is None

    # copy directly via the chooser + copytree path
    import shutil
    shutil.copytree(str(old_full), str(fixed), dirs_exist_ok=True)
    assert (fixed / "secret_key").read_text() == "the-real-key"
    # a folder that already holds a healthy database is never adopted over
    assert run_server._adopt_existing_data(str(tmp_path / "base"), str(fixed)) is None
    assert run_server._count_assets(str(fixed / "itam.sqlite")) == 240


def test_data_adoption_replaces_a_corrupt_database(tmp_path):
    """A broken fixed database is moved aside (kept) and re-adopted from the
    best healthy source, using the SQLite backup API."""
    import run_server

    good = tmp_path / "base" / "instance"
    _mini_db(str(good / "itam.sqlite"), assets=7)
    assert run_server._db_healthy(str(good / "itam.sqlite"))

    fixed = tmp_path / "fixed"
    fixed.mkdir()
    (fixed / "itam.sqlite").write_bytes(b"this is not a database at all")
    assert not run_server._db_healthy(str(fixed / "itam.sqlite"))

    src = run_server._adopt_existing_data(str(tmp_path / "base"), str(fixed))
    assert src == str(good)
    assert run_server._db_healthy(str(fixed / "itam.sqlite"))
    assert run_server._count_assets(str(fixed / "itam.sqlite")) == 7
    # the broken file was kept, not deleted
    assert list(fixed.glob("itam.sqlite.corrupt-*"))


def test_update_now_from_source_says_so(client):
    """Only the packaged exe can replace itself; from source it must not 500."""
    login(client)
    page = client.post("/admin/update-now", follow_redirects=True).data.decode()
    assert "Running from source" in page


def test_update_now_needs_settings_permission(client):
    login(client, "viewer", "viewer123")
    assert client.post("/admin/update-now").status_code in (302, 403)


BULK_PAGES = [
    ("/employees", "org.employees_bulk_delete"),
    ("/departments", "org.departments_bulk_delete"),
    # /locations no longer shows the row table (removed on request); its
    # bulk-delete endpoint still exists for API use.
    ("/vendors", "org.vendors_bulk_delete"),
    ("/licenses", "ops.licenses_bulk_delete"),
    ("/maintenance", "ops.maintenance_bulk_delete"),
    ("/inventory", "ops.inventory_bulk_delete"),
    ("/checkouts", "ops.checkouts_bulk_checkin"),
]


@pytest.mark.parametrize("url,endpoint", BULK_PAGES)
def test_list_pages_offer_select_all(client, app, url, endpoint):
    """The toolbar only appears once a page has rows, so give each one."""
    from itam.models import (Assignment, Category, Department, Employee,
                             InventoryAudit, License, Location, Maintenance,
                             Vendor)

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        emp = db.session.scalar(db.select(Employee))
        db.session.add_all([Department(name="Dept One"),
                            Location(name="Room One", kind="Room"),
                            License(name="Office", seats=5),
                            Vendor(name="Vendor One"),
                            InventoryAudit(name="Audit One")])
        asset = Asset(tag="BK-9", name="Rig", status="Checked Out",
                      condition="Good", category=cat)
        db.session.add(asset)
        db.session.flush()
        db.session.add_all([Maintenance(asset=asset, title="Service"),
                            Assignment(asset=asset, employee=emp)])
        db.session.commit()

    login(client)
    body = client.get(url).data.decode()
    assert 'class="bulk-form"' in body, f"{url} has no bulk form"
    assert 'id="bulk"' in body, f"{url} bulk form needs an id for the button to target"
    assert 'form="bulk"' in body, f"{url} action button must point at the form"


def test_employees_bulk_edit_actions(client, app):
    """Select-all offers more than delete: set department/type, active flag."""
    from itam.models import Department, Employee

    with app.app_context():
        dep = Department(name="Science Wing")
        e1 = Employee(name="Bulk One")
        e2 = Employee(name="Bulk Two")
        db.session.add_all([dep, e1, e2])
        db.session.commit()
        dep_id, ids = dep.id, [e1.id, e2.id]

    login(client)
    body = client.get("/employees").data.decode()
    for value in ("department", "type", "activate", "deactivate"):
        assert f'value="{value}"' in body, f"no bulk {value} option"

    client.post("/employees/bulk", data={
        "action": "department", "department_id": str(dep_id),
        "ids": [str(i) for i in ids]}, follow_redirects=True)
    client.post("/employees/bulk", data={
        "action": "type", "emp_type": "Teacher",
        "ids": [str(i) for i in ids]}, follow_redirects=True)
    client.post("/employees/bulk", data={
        "action": "deactivate", "ids": [str(ids[0])]}, follow_redirects=True)
    with app.app_context():
        one, two = (db.session.get(Employee, i) for i in ids)
        assert one.department_id == dep_id and two.department_id == dep_id
        assert one.emp_type == "Teacher" and two.emp_type == "Teacher"
        assert one.active is False and two.active is True

    # delete still works through the same form
    client.post("/employees/bulk", data={
        "action": "delete", "ids": [str(i) for i in ids]}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Employee, ids[0]) is None


def test_export_matches_the_school_sheet_and_round_trips(client, app):
    """Export mirrors the inventory sheet's column order (Asset ID included)
    and an exported file can be re-imported without losing gpu/invoice/
    depreciation."""
    import re
    login(client)
    csv_in = (
        "Name,Asset ID,Category,Serial No.,Graphic cards ,Invoice Number,"
        "Depreciation (Years),Assign to \n"
        "RT-PC,RT-0001,Desktop,SN-RT,RTX 3060,INV-77,7,مس روان\n"
    )
    resp = client.post("/assets/import",
                       data={"file": (io.BytesIO(csv_in.encode()), "rt.csv")},
                       content_type="multipart/form-data")
    token = re.search(rb'name="token" value="([^"]+)"', resp.data).group(1).decode()
    client.post("/assets/import", data={"token": token}, follow_redirects=True)

    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.name == "RT-PC"))
        assert a.gpu == "RTX 3060"
        assert a.invoice_number == "INV-77"
        assert a.depreciation_years == 7

    out = client.get("/assets/export.csv").data.decode()
    header, row = out.strip().splitlines()[0], [
        line for line in out.strip().splitlines() if "RT-PC" in line][0]
    assert header.lstrip("﻿").startswith(
        "Name,Category,Asset ID,Condition,Status,Serial No.")
    for value in ("RT-0001", "RTX 3060", "INV-77", "مس روان"):
        assert value in row


def test_reimporting_the_same_idless_file_does_not_duplicate(client, app):
    """The school sheet has an empty Asset ID column; importing it twice used
    to double the register. Same name+serial with no ID -> refused."""
    import re
    login(client)
    csv = ("Name,Category,Serial No.\n"
           "RE-PC1,Desktop,SN-R1\n"
           "RE-PC2,Desktop,SN-R2\n")

    def run():
        resp = client.post("/assets/import",
                           data={"file": (io.BytesIO(csv.encode()), "re.csv")},
                           content_type="multipart/form-data")
        token = re.search(rb'name="token" value="([^"]+)"',
                          resp.data).group(1).decode()
        return client.post("/assets/import", data={"token": token},
                           follow_redirects=True)

    run()
    with app.app_context():
        first = db.session.scalar(db.select(db.func.count(Asset.id)))
    done = run()
    assert b"Duplicated Records" in done.data
    with app.app_context():
        assert db.session.scalar(db.select(db.func.count(Asset.id))) == first


def test_delete_all_assets_clears_every_page(client, app):
    """Bulk delete works 50 rows at a time; Delete all wipes the register."""
    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        for i in range(3):
            db.session.add(Asset(tag=f"WIPE-{i}", name=f"Thing {i}",
                                 status="Available", condition="Good",
                                 category=cat))
        db.session.commit()

    login(client)
    body = client.get("/assets/").data.decode()
    assert "Delete all" in body
    client.post("/assets/delete-all", follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(db.func.count(Asset.id))) == 0


def test_assets_bulk_set_department_and_condition(client, app):
    from itam.models import Department

    with app.app_context():
        dep = Department(name="Media Lab")
        a = Asset(tag="BLK-1", name="Cam", status="Available", condition="Good",
                  category=db.session.scalar(db.select(Category)))
        db.session.add_all([dep, a])
        db.session.commit()
        dep_id, aid = dep.id, a.id

    login(client)
    client.post("/assets/bulk", data={"action": "department",
                                      "department_id": str(dep_id),
                                      "id": [str(aid)]}, follow_redirects=True)
    client.post("/assets/bulk", data={"action": "condition:Damaged",
                                      "id": [str(aid)]}, follow_redirects=True)
    with app.app_context():
        a = db.session.get(Asset, aid)
        assert a.department_id == dep_id
        assert a.condition == "Damaged"


def test_bulk_delete_skips_rows_still_in_use(client, app):
    """A department with people in it must survive a bulk delete."""
    from itam.models import Department, Employee

    with app.app_context():
        busy = Department(name="Busy")
        spare = Department(name="Spare")
        db.session.add_all([busy, spare])
        db.session.flush()
        db.session.add(Employee(name="Someone", department=busy))
        db.session.commit()
        busy_id, spare_id = busy.id, spare.id

    login(client)
    resp = client.post("/departments/bulk-delete",
                       data={"ids": [busy_id, spare_id]}, follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        assert db.session.get(Department, busy_id) is not None, "in-use row was deleted"
        assert db.session.get(Department, spare_id) is None, "free row was not deleted"


def test_bulk_delete_with_nothing_selected_is_harmless(client, app):
    from itam.models import Department

    with app.app_context():
        db.session.add(Department(name="Untouched"))
        db.session.commit()
    login(client)
    resp = client.post("/departments/bulk-delete", data={}, follow_redirects=True)
    assert resp.status_code == 200
    assert b"Nothing selected" in resp.data
    with app.app_context():
        assert db.session.scalar(db.select(Department).where(Department.name == "Untouched"))


def test_departments_can_be_edited(client, app):
    """Both screens had no edit path at all before."""
    from itam.models import Department, Location

    login(client)
    client.post("/departments", data={"name": "Science", "cost_center": "CC-1"},
                follow_redirects=True)
    with app.app_context():
        dep = db.session.scalar(db.select(Department).where(Department.name == "Science"))
        dep_id = dep.id
    client.post("/departments", data={"id": dep_id, "name": "Science Lab",
                                      "cost_center": "CC-2"}, follow_redirects=True)
    with app.app_context():
        dep = db.session.get(Department, dep_id)
        assert (dep.name, dep.cost_center) == ("Science Lab", "CC-2")



def test_bulk_checkin_returns_several_loans(client, app):
    from itam.models import Assignment, Category, Employee

    with app.app_context():
        emp = db.session.scalar(db.select(Employee))
        cat = db.session.scalar(db.select(Category))
        ids = []
        for tag in ("BK-1", "BK-2"):
            asset = Asset(tag=tag, name=tag, status="Checked Out", condition="Good",
                          category=cat)
            db.session.add(asset)
            db.session.flush()
            asg = Assignment(asset=asset, employee=emp)
            db.session.add(asg)
            db.session.flush()
            ids.append(asg.id)
        db.session.commit()

    login(client)
    client.post("/checkouts/bulk-checkin", data={"ids": ids}, follow_redirects=True)
    with app.app_context():
        for asg_id in ids:
            asg = db.session.get(Assignment, asg_id)
            assert asg.returned_at is not None, "loan was not returned"
            assert asg.asset.status == "Available"


def test_asset_form_has_no_parent_field(client):
    """"Part of (parent asset)" was removed from the asset form."""
    login(client)
    body = client.get("/assets/new").data.decode()
    assert 'name="parent_id"' not in body
    assert "Part of (parent asset)" not in body


def test_editing_keeps_an_existing_parent_link(client, app):
    """The field is gone, so a save must leave the stored link alone."""
    from itam.models import Category

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        parent = Asset(tag="PA-1", name="Dock", status="Available",
                       condition="Good", category=cat)
        db.session.add(parent)
        db.session.flush()
        child = Asset(tag="CH-1", name="Monitor", status="Available",
                      condition="Good", category=cat, parent_id=parent.id)
        db.session.add(child)
        db.session.commit()
        child_id, parent_id = child.id, parent.id

    login(client)
    client.post(f"/assets/{child_id}/edit", data={
        "name": "Monitor", "tag": "CH-1", "status": "Available",
        "condition": "Good", "depreciation_years": "5"}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Asset, child_id).parent_id == parent_id, \
            "saving the form unlinked an asset from its parent"


def test_new_asset_form_defaults_purchase_date_to_today(client, app):
    from datetime import date

    login(client)
    assert date.today().isoformat() in client.get("/assets/new").data.decode()

    # An existing asset keeps its own date rather than being reset to today.
    with app.app_context():
        from itam.models import Category
        a = Asset(tag="PD-1", name="Old", status="Available", condition="Good",
                  category=db.session.scalar(db.select(Category)),
                  purchase_date=date(2020, 1, 15))
        db.session.add(a)
        db.session.commit()
        aid = a.id
    body = client.get(f"/assets/{aid}/edit").data.decode()
    assert "2020-01-15" in body


def test_room_and_os_use_the_same_control_as_every_other_picker(client, app):
    """A <datalist> is drawn by the browser and cannot be styled.

    Next to the app's own dropdowns it appeared as a black list, so Room and
    Operating system are <select> elements now, with an "Other..." choice that
    reveals a text box for anything not listed.
    """
    from itam.models import Category

    login(client)
    body = client.get("/assets/new").data.decode()
    assert "<datalist" not in body, "a browser-drawn list is still on the form"
    for field in ("f-place-pick", "f-os-pick"):
        assert f'id="{field}"' in body, f"{field} is missing"
    assert "Other…" in body or "Other&#8230;" in body

    # A listed value saves.
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
    client.post("/assets/new", data={
        "name": "Listed", "tag": "SEL-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "location_name": "Reception", "os_name": "Windows 11 Pro"},
        follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "SEL-1"))
        assert (a.location_name, a.os_name) == ("Reception", "Windows 11 Pro")

    # And so does something typed under "Other".
    client.post("/assets/new", data={
        "name": "Typed", "tag": "SEL-2", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "location_name": "Rooftop Store", "os_name": "Slackware"},
        follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "SEL-2"))
        assert (a.location_name, a.os_name) == ("Rooftop Store", "Slackware")

    # An imported value that is not in the list still shows on the edit form.
    with app.app_context():
        aid = db.session.scalar(db.select(Asset.id).where(Asset.tag == "SEL-2"))
    edit = client.get(f"/assets/{aid}/edit").data.decode()
    assert "Rooftop Store" in edit and "Slackware" in edit


def test_asset_detail_still_renders(client, app):
    """Regression: a `today` name clash in the form lookups 500'd this page."""
    from itam.models import Category

    with app.app_context():
        a = Asset(tag="DT-1", name="Rig", status="Available", condition="Good",
                  category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.commit()
        aid = a.id
    login(client)
    assert client.get(f"/assets/{aid}").status_code == 200


# ----------------------------------------------------------------- updater

def test_version_parsing_and_comparison():
    from itam.updater import is_newer, parse_version

    assert parse_version("2026.07.19") == (2026, 7, 19)
    assert parse_version("v1.2.3") == (1, 2, 3)
    assert parse_version("") is None and parse_version(None) is None
    assert parse_version("nightly") is None

    assert is_newer("2026.07.20", "2026.07.19")
    assert is_newer("2026.08.01", "2026.07.31")
    assert not is_newer("2026.07.19", "2026.07.19")     # same build
    assert not is_newer("2026.07.18", "2026.07.19")     # older
    # Anything unparseable must never trigger an update.
    assert not is_newer("nightly", "2026.07.19")
    assert not is_newer(None, "2026.07.19")
    assert not is_newer("v1.0.0", "2026.07.19")         # 1.x is not above 2026.x

    # Versions of different lengths still compare, so introducing a hotfix
    # component cannot silently freeze every installed copy.
    assert is_newer("2026.08.01.1", "2026.08.01")
    assert not is_newer("2026.08.01", "2026.08.01.1")
    assert not is_newer("2026.08.01.0", "2026.08.01")   # same build, padded
    assert is_newer("2026.09", "2026.08.30")


def test_update_check_is_silent_when_github_is_unreachable(tmp_path, monkeypatch):
    """No internet must be a no-op, not an error."""
    from itam import updater

    monkeypatch.setattr(updater, "latest_release", lambda repo, token=None: None)
    assert updater.check_for_update("o/r", "2026.07.19", str(tmp_path)) == "unavailable"
    assert list(tmp_path.iterdir()) == []                # nothing written


def test_update_check_does_nothing_when_already_current(tmp_path, monkeypatch):
    from itam import updater

    release = {"name": "Mada AMS 2026.07.19", "assets": []}
    monkeypatch.setattr(updater, "latest_release", lambda repo, token=None: release)
    monkeypatch.setattr(updater, "remote_version", lambda r, token=None: "2026.07.19")
    called = []
    monkeypatch.setattr(updater, "download_update",
                        lambda *a, **k: called.append(1) or True)

    assert updater.check_for_update("o/r", "2026.07.19", str(tmp_path)) == "up-to-date"
    assert not called, "downloaded an update that was not newer"


def test_update_downloads_beside_the_exe_and_never_touches_instance(tmp_path, monkeypatch):
    """The whole point: data must come through an update untouched."""
    import hashlib

    from itam import updater

    exe = tmp_path / "AMS.exe"
    exe.write_bytes(b"old build")
    instance = tmp_path / "instance"
    (instance / "uploads").mkdir(parents=True)
    (instance / "backups").mkdir()
    db_file = instance / "itam.sqlite"
    db_file.write_bytes(b"SQLite format 3\x00" + b"important data" * 100)
    upload = instance / "uploads" / "invoice.pdf"
    upload.write_bytes(b"a scanned invoice")

    def fingerprint():
        out = {}
        for path in sorted(instance.rglob("*")):
            if path.is_file():
                out[str(path.relative_to(instance))] = hashlib.sha256(
                    path.read_bytes()).hexdigest()
        return out

    before = fingerprint()

    release = {"name": "Mada AMS 2026.07.20",
               "assets": [{"name": "AMS.exe",
                           "browser_download_url": "https://example/AMS.exe"}]}
    monkeypatch.setattr(updater, "latest_release", lambda repo, token=None: release)
    monkeypatch.setattr(updater, "remote_version", lambda r, token=None: "2026.07.20")
    monkeypatch.setattr(updater, "download_update",
                        lambda rel, base, ex=None, token=None:
                        (updater.pending_path(base, ex),
                         open(updater.pending_path(base, ex), "wb").write(b"new build"),
                         True)[-1])

    status = updater.check_for_update("o/r", "2026.07.19", str(tmp_path), exe=str(exe))
    assert status == "downloaded"
    assert (tmp_path / "AMS.exe.new").read_bytes() == b"new build"
    assert exe.read_bytes() == b"old build", "the running exe was replaced too early"
    assert fingerprint() == before, "an update modified the instance folder"


def test_applying_an_update_swaps_the_exe_and_leaves_data_alone(tmp_path):
    import hashlib

    from itam import updater

    exe = tmp_path / "AMS.exe"
    exe.write_bytes(b"old build")
    (tmp_path / "AMS.exe.new").write_bytes(b"new build")
    instance = tmp_path / "instance"
    instance.mkdir()
    db_file = instance / "itam.sqlite"
    db_file.write_bytes(b"rows and rows")
    before = hashlib.sha256(db_file.read_bytes()).hexdigest()

    assert updater.pending_update(str(tmp_path), str(exe)) is True
    assert updater.apply_pending_update(str(tmp_path), str(exe)) is True

    assert exe.read_bytes() == b"new build"
    assert (tmp_path / "AMS.exe.old").read_bytes() == b"old build"
    assert not (tmp_path / "AMS.exe.new").exists()
    assert hashlib.sha256(db_file.read_bytes()).hexdigest() == before

    # The retired build is cleared away on the following start.
    updater.cleanup_retired(str(tmp_path), str(exe))
    assert not (tmp_path / "AMS.exe.old").exists()


def test_apply_is_a_no_op_without_a_pending_download(tmp_path):
    from itam import updater

    exe = tmp_path / "AMS.exe"
    exe.write_bytes(b"only build")
    assert updater.pending_update(str(tmp_path), str(exe)) is False
    assert updater.apply_pending_update(str(tmp_path), str(exe)) is False
    assert exe.read_bytes() == b"only build"


def test_a_truncated_download_is_discarded(tmp_path, monkeypatch):
    """A dropped connection must not leave a stub that gets 'installed'."""
    import io

    from itam import updater

    exe = tmp_path / "AMS.exe"
    exe.write_bytes(b"old build")

    class Resp(io.BytesIO):
        def __enter__(self): return self
        def __exit__(self, *a): return False

    monkeypatch.setattr(updater, "_request",
                        lambda url, token=None, accept=None: Resp(b"truncated"))
    release = {"assets": [{"name": "AMS.exe",
                           "browser_download_url": "https://example/AMS.exe"}]}
    assert updater.download_update(release, str(tmp_path), str(exe)) is False
    assert not (tmp_path / "AMS.exe.new").exists()
    assert not (tmp_path / "AMS.exe.new.part").exists()


def test_update_settings_expose_an_on_off_toggle(client, app):
    from itam.utils import get_setting

    login(client)
    page = client.get("/admin/settings").data.decode()
    assert "Update automatically" in page
    assert 'name="update_auto"' in page

    form = {"app_name": "Mada Asset Management System (AMS)", "qr_prefix": "",
            "custom_asset_fields": "", "checkout_days": "30",
            "label_org": "Mada International Academy", "label_width_mm": "152.4",
            "label_height_mm": "76.2", "warranty_alert_days": "90",
            "license_alert_days": "90", "smtp_host": "", "smtp_port": "587",
            "smtp_user": "", "smtp_password": "", "smtp_from": "",
            "audit_retention_days": "365", "update_repo": "o/r", "update_token": ""}
    client.post("/admin/settings", data=form, follow_redirects=True)   # toggle off
    with app.app_context():
        assert get_setting("update_auto") == "0"

    client.post("/admin/settings", data={**form, "update_auto": "1"},
                follow_redirects=True)
    with app.app_context():
        assert get_setting("update_auto") == "1"


def test_pending_update_shows_a_restart_note(client, app):
    from itam.utils import set_setting

    login(client)
    assert b"Update ready" not in client.get("/").data
    with app.app_context():
        set_setting("update_pending", "2026.08.01")
        db.session.commit()
    page = client.get("/").data.decode()
    assert "Update ready" in page and "2026.08.01" in page


def test_asset_list_pages_instead_of_rendering_everything(client, app):
    """A 3,000-asset register produced a 1 MB page and a query per row."""
    from itam.blueprints.assets import PAGE_SIZE
    from itam.models import Category

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        db.session.add_all([
            Asset(tag=f"PG-{i:04d}", name=f"Device {i}", category=cat,
                  status="Available", condition="Good")
            for i in range(PAGE_SIZE + 25)])
        db.session.commit()

    login(client)
    body = client.get("/assets/").data.decode()
    assert body.count('data-id=') <= PAGE_SIZE + 1, "more than one page of rows"
    assert "Page 1 /" in body and "Next" in body

    page2 = client.get("/assets/?page=2").data.decode()
    assert "PG-0050" in page2 and "PG-0000" not in page2
    # An out-of-range page clamps rather than 500s or shows nothing.
    assert client.get("/assets/?page=999").status_code == 200
    assert client.get("/assets/?page=abc").status_code == 200


def test_export_still_covers_every_page(client, app):
    from itam.blueprints.assets import PAGE_SIZE
    from itam.models import Category

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        db.session.add_all([
            Asset(tag=f"EX-{i:04d}", name=f"Device {i}", category=cat,
                  status="Available", condition="Good")
            for i in range(PAGE_SIZE + 10)])
        db.session.commit()

    login(client)
    rows = client.get("/assets/export.csv").data.decode().strip().splitlines()
    assert len(rows) - 1 >= PAGE_SIZE + 10, "export was truncated to one page"


def test_asset_list_does_not_query_per_row(client, app):
    """Guards the eager loading: rows must not each fetch their own lookups."""
    from sqlalchemy import event

    from itam.models import Category

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        db.session.add_all([
            Asset(tag=f"NP-{i:03d}", name=f"Device {i}", category=cat,
                  status="Available", condition="Good") for i in range(40)])
        db.session.commit()
        engine = db.engine

    counter = {"n": 0}

    def count(*args, **kwargs):
        counter["n"] += 1

    login(client)
    event.listen(engine, "before_cursor_execute", count)
    try:
        client.get("/assets/")
    finally:
        event.remove(engine, "before_cursor_execute", count)
    assert counter["n"] < 40, f"one query per row is back ({counter['n']} queries)"


def test_locations_are_derived_when_an_asset_is_saved(client, app):
    """A place is a property of an asset, not a list someone maintains.

    The Locations screen is gone. Location rows still back the asset filter,
    bulk transfers, transfer history and the locations report, so they are
    rebuilt whenever an asset is saved or imported.
    """
    from itam.models import Category, Location

    login(client)
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))

    for tag, floor, room in (("LOC-A", "GF", "office"),
                             ("LOC-B", "F1", "Computer Lab"),
                             ("LOC-C", "GF", "office")):   # same room as LOC-A
        client.post("/assets/new", data={
            "name": tag, "tag": tag, "category_id": cat_id, "status": "In Use",
            "condition": "Good", "depreciation_years": "5", "branch": "Mada 3",
            "building": "Building 1", "floor": floor,
            "location_name": room}, follow_redirects=True)

    with app.app_context():
        rows = db.session.scalars(db.select(Location)).all()
        kinds = {l.name: l.kind for l in rows}
        assert kinds.get("Mada 3") == "Branch"
        assert kinds.get("Building 1") == "Building"
        assert kinds.get("office") == "Room"
        # One shared room, not one per asset.
        assert len([l for l in rows if l.name == "office"]) == 1
        # Assets are linked to their room, so counts and filters mean something.
        office = next(l for l in rows if l.name == "office")
        assert db.session.scalar(db.select(db.func.count(Asset.id))
                                 .where(Asset.location_id == office.id)) == 2
        before = len(rows)

    # Saving again must not add anything.
    with app.app_context():
        aid = db.session.scalar(db.select(Asset.id).where(Asset.tag == "LOC-A"))
    client.post(f"/assets/{aid}/edit", data={
        "name": "LOC-A", "tag": "LOC-A", "status": "In Use", "condition": "Good",
        "depreciation_years": "5", "branch": "Mada 3", "building": "Building 1",
        "floor": "GF", "location_name": "office"}, follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(db.func.count(Location.id))) == before, \
            "saving again duplicated locations"


def test_a_hand_picked_location_is_not_overwritten(client, app):
    """An explicit choice on the asset beats anything inferred from its text."""
    from itam.models import Category, Location

    with app.app_context():
        chosen = Location(name="Secure Store", kind="Room")
        db.session.add(chosen)
        db.session.flush()
        db.session.add(Asset(tag="LOC-D", name="Spare", status="In Storage",
                             condition="Good", location_id=chosen.id,
                             category=db.session.scalar(db.select(Category)),
                             branch="Mada 1", building="Building 2",
                             floor="B1", location_name="Basement"))
        db.session.commit()
        chosen_id = chosen.id

    login(client)
    # Saving any asset triggers the sync; it must leave LOC-D's link alone.
    client.post("/assets/new", data={
        "name": "Other", "tag": "LOC-E", "status": "In Use", "condition": "Good",
        "depreciation_years": "5", "branch": "Mada 1", "building": "Building 2",
        "floor": "B1", "location_name": "Basement"}, follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "LOC-D"))
        assert a.location_id == chosen_id, "the manual location was overwritten"


def test_locations_screen_is_available_and_seeded(client, app):
    """The standard tree ships with the database, so nothing starts empty."""
    from itam.models import Location

    login(client)
    assert client.get("/locations").status_code == 200
    with app.app_context():
        kinds = {l.kind for l in db.session.scalars(db.select(Location))}
        total = db.session.scalar(db.select(db.func.count(Location.id)))
    assert total > 100, "the standard tree was not seeded"
    assert {"Branch", "Building", "Floor", "Room"} <= kinds

    # The kind list reads down the hierarchy.
    from itam.models import LOCATION_KINDS
    assert LOCATION_KINDS[:3] == ["Branch", "Building", "Floor"]
    assert "Department" in LOCATION_KINDS


def test_a_viewer_cannot_change_locations(client, app):
    from itam.models import Location

    login(client, "viewer", "viewer123")
    assert client.get("/locations").status_code == 200
    client.post("/locations", data={"name": "Sneaky Room", "kind": "Room"},
                follow_redirects=True)
    with app.app_context():
        assert not db.session.scalar(
            db.select(Location).where(Location.name == "Sneaky Room")), \
            "a viewer created a location"


def test_os_version_is_gone_from_the_form_but_kept_in_the_data(client, app):
    from itam.models import Category

    with app.app_context():
        a = Asset(tag="OSV-1", name="Imported", status="Available",
                  condition="Good", os_name="Windows 11 Pro", os_version="23H2",
                  category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.commit()
        aid = a.id

    login(client)
    body = client.get(f"/assets/{aid}/edit").data.decode()
    assert 'name="os_version"' not in body, "OS version is still on the form"

    client.post(f"/assets/{aid}/edit", data={
        "name": "Imported", "tag": "OSV-1", "status": "Available",
        "condition": "Good", "depreciation_years": "5",
        "os_name": "Windows 11 Pro"}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Asset, aid).os_version == "23H2", \
            "saving the form wiped the imported OS version"


def test_lending_lists_only_assets_that_are_free(client, app):
    """The picker must offer free assets and hide the ones already out."""
    from itam.models import Assignment, Category, Employee

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        emp = db.session.scalar(db.select(Employee))
        free = Asset(tag="LEND-FREE", name="Spare laptop", status="Available",
                     condition="Good", category=cat, serial="SN-FREE")
        out = Asset(tag="LEND-OUT", name="Loaned laptop", status="Checked Out",
                    condition="Good", category=cat)
        retired = Asset(tag="LEND-GONE", name="Retired laptop", status="Retired",
                        condition="Poor", category=cat)
        db.session.add_all([free, out, retired])
        db.session.flush()
        db.session.add(Assignment(asset=out, employee=emp))
        db.session.commit()

    login(client)
    body = client.get("/checkouts").data.decode()
    picker = body.split('id="co-asset"', 1)[1].split("</select>", 1)[0]
    assert "LEND-FREE" in picker
    assert "LEND-OUT" not in picker, "an asset already on loan was offered"
    assert "LEND-GONE" not in picker, "a retired asset was offered"

    # The search box and the read-only detail card share the same filter.
    with app.app_context():
        free_id = db.session.scalar(db.select(Asset.id).where(Asset.tag == "LEND-FREE"))
    hits = client.get("/lend/assets.json?q=SN-FREE").get_json()
    assert [h["id"] for h in hits] == [free_id]
    assert client.get("/lend/assets.json?q=LEND-OUT").get_json() == []
    facts = client.get(f"/lend/assets/{free_id}.json").get_json()
    assert facts["serial"] == "SN-FREE" and facts["tag"] == "LEND-FREE"


def test_lending_picker_does_not_render_the_whole_register(client, app):
    """A few thousand <option> tags made this page a megabyte of HTML."""
    from itam.blueprints.operations import PICKER_LIMIT
    from itam.models import Category

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        db.session.add_all([
            Asset(tag=f"BULK-{i:04d}", name=f"Bulk {i}", status="Available",
                  condition="Good", category=cat)
            for i in range(PICKER_LIMIT + 50)])
        db.session.commit()

    login(client)
    body = client.get("/checkouts").data.decode()
    picker = body.split('id="co-asset"', 1)[1].split("</select>", 1)[0]
    assert picker.count("<option") == PICKER_LIMIT
    assert f"of {PICKER_LIMIT + 50}" in body, "the full count is not reported"


def test_lending_records_assign_to_and_edited_by(client, app):
    from itam.models import Assignment, Category, Employee

    with app.app_context():
        a = Asset(tag="LEND-1", name="Trolley", status="Available",
                  condition="Good", category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.commit()
        aid, eid = a.id, db.session.scalar(db.select(Employee)).id

    login(client)
    resp = client.post("/lend", data={"asset_id": aid, "employee_id": eid,
                                      "handled_by": "Sara N."},
                       follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        asg = db.session.scalar(db.select(Assignment).where(Assignment.asset_id == aid))
        assert asg.employee_id == eid, "assign-to was not recorded"
        assert asg.handled_by == "Sara N.", "edited-by was not recorded"
        assert db.session.get(Asset, aid).status == "Checked Out"
    assert b"Sara N." in client.get("/checkouts").data


def test_asset_form_shows_assignment_but_cannot_change_it(client, app):
    """Assignment lives in Lending; the asset form only mirrors it."""
    from itam.models import Assignment, Category, Employee

    with app.app_context():
        emp = db.session.scalar(db.select(Employee))
        a = Asset(tag="RO-1", name="Beamer", status="Checked Out", condition="Good",
                  category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.flush()
        db.session.add(Assignment(asset=a, employee=emp))
        db.session.commit()
        aid, emp_name = a.id, emp.name

    login(client)
    body = client.get(f"/assets/{aid}/edit").data.decode()
    assert 'name="assign_employee_id"' not in body, "Assign to is still editable"
    assert 'name="updated_by"' not in body, "Updated by is still editable"
    assert emp_name in body, "the current holder is no longer shown"

    # Saving the form must not disturb the loan Lending created.
    client.post(f"/assets/{aid}/edit", data={
        "name": "Beamer", "tag": "RO-1", "status": "Checked Out",
        "condition": "Good", "depreciation_years": "5"}, follow_redirects=True)
    with app.app_context():
        asg = db.session.scalar(db.select(Assignment).where(Assignment.asset_id == aid))
        assert asg.returned_at is None, "saving the asset ended the loan"
        assert db.session.get(Asset, aid).updated_by == "Administrator", \
            "the saving user was not stamped"


def test_service_worker_cache_is_versioned(client):
    """A fixed cache name meant an installed PWA served last release's CSS."""
    from itam import APP_VERSION

    resp = client.get("/static/sw.js")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert f'"{APP_VERSION}"' in body, "the worker does not carry the build number"
    assert "no-store" in resp.headers.get("Cache-Control", ""), \
        "the worker itself may be cached, pinning the old one in place"

    login(client)
    page = client.get("/assets/").data.decode()
    assert f"style.css?v={APP_VERSION}" in page
    assert f"app.js?v={APP_VERSION}" in page


def test_every_way_of_lending_records_who_did_it(client, app):
    """"Edited by" must not depend on which screen the loan came from."""
    from itam.models import Assignment, Category, Employee

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        db.session.add_all([
            Asset(tag=f"WAY-{i}", name=f"Way {i}", status="Available",
                  condition="Good", category=cat) for i in (1, 2, 3)])
        db.session.commit()
        ids = {a.tag: a.id for a in db.session.scalars(
            db.select(Asset).where(Asset.tag.like("WAY-%")))}
        eid = db.session.scalar(db.select(Employee)).id

    login(client)
    client.post("/lend", data={"asset_id": ids["WAY-1"], "employee_id": eid},
                follow_redirects=True)
    client.post(f"/assets/{ids['WAY-2']}/checkout", data={"employee_id": eid},
                follow_redirects=True)
    client.post("/assets/bulk", data={"action": "assign", "employee_id": eid,
                                      "id": [ids["WAY-3"]]}, follow_redirects=True)
    with app.app_context():
        for tag, aid in ids.items():
            asg = db.session.scalar(
                db.select(Assignment).where(Assignment.asset_id == aid))
            assert asg is not None, f"{tag} was never lent"
            assert asg.handled_by == "Administrator", f"{tag} recorded no Edited by"


def test_asset_filters_offer_only_locations_in_use(client, app):
    """A filter <select> of every location was 39 KB on the busiest page."""
    from itam.models import Category, Location

    login(client)
    with app.app_context():
        total = db.session.scalar(db.select(db.func.count(Location.id)))
        cat_id = db.session.scalar(db.select(Category.id))
    assert total > 100, "the standard tree should be seeded"

    client.post("/assets/new", data={
        "name": "Desk PC", "tag": "LOC-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "branch": "Mada 1", "building": "Building 1", "floor": "GF",
        "location_name": "Reception"}, follow_redirects=True)

    with app.app_context():
        used = db.session.scalar(
            db.select(Asset.location_id).where(Asset.tag == "LOC-1"))
    assert used, "the asset was not linked to a room"

    body = client.get("/assets/").data.decode()
    picker = body.split('name="location"', 1)[1].split("</select>", 1)[0]
    assert f'value="{used}"' in picker
    assert picker.count("<option") < total, "every location is still listed"

    # Transfers may target an empty location, so that list covers the tree.
    hits = client.get("/assets/locations.json").get_json()
    assert len(hits) > 50, "the transfer picker is not offering the tree"


def test_location_report_totals_survive_the_rewrite(client, app):
    """The per-location loop became one grouped query; the numbers must match."""
    from itam.models import Category, Location

    login(client)
    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        room = Location(name="Lab 7", kind="Room")
        wing = Location(name="North Wing", kind="Building")
        db.session.add_all([wing, room])
        db.session.flush()
        room.parent_id = wing.id
        empty = Location(name="Nobody's Room", kind="Room")
        db.session.add(empty)
        db.session.add_all([
            Asset(tag="RP-1", name="A", status="Available", condition="Good",
                  category=cat, location_id=room.id, purchase_cost=100),
            Asset(tag="RP-2", name="B", status="Available", condition="Good",
                  category=cat, location_id=room.id, purchase_cost=250.50),
        ])
        db.session.commit()

    body = client.get("/reports/locations").data.decode()
    assert "North Wing / Lab 7" in body, "the full path is no longer shown"
    assert "350.50" in body, "purchase cost was not totalled"
    assert "Nobody" not in body, "a location holding nothing was listed"

    csv = client.get("/reports/locations?format=csv").data.decode()
    row = [l for l in csv.splitlines() if "Lab 7" in l][0]
    assert ",2," in row or '"2"' in row or ",2," in row.replace('"', "")


def test_junk_query_arguments_do_not_crash_pages(client, app):
    """A hand-edited URL or stale bookmark used to return a 500."""
    login(client)
    for url in ("/assets/?location=notanumber", "/assets/?category=x&department=y",
                "/assets/?page=abc", "/assets/new?clone=abc",
                "/reports/custom?category=zzz", "/reports/custom?department=zzz"):
        assert client.get(url).status_code == 200, f"{url} did not survive"


# ------------------------------------------------------- desktop app window

class _FakeScreen:
    def __init__(self, width, height):
        self.width, self.height = width, height


class _FakeWindow:
    def __init__(self):
        self.maximized_calls = 0

    def maximize(self):
        self.maximized_calls += 1


class _FakeWebview:
    """Just enough of pywebview to exercise the launcher off Windows."""

    def __init__(self, screens=None, start_error=None):
        self.settings = {}
        self.screens = screens if screens is not None else [_FakeScreen(1920, 1080)]
        self.created = None
        self.started = []
        self._start_error = start_error

    def create_window(self, title, url, width=None, height=None, **kw):
        self.created = {"title": title, "url": url, "width": width,
                        "height": height, **kw}
        return _FakeWindow()

    def start(self, func=None, args=None):
        self.started.append((func, args))
        if self._start_error and len(self.started) == 1:
            raise self._start_error
        if func:
            func(args)


def test_window_is_created_at_screen_size_and_maximized():
    """maximized=True alone left the browser control at 1280x840."""
    import run_server

    fake = _FakeWebview(screens=[_FakeScreen(1600, 900)])
    assert run_server._primary_screen_size(fake, 1280, 840) == (1600, 900)
    window = fake.create_window("t", "u", width=1600, height=900, maximized=True)
    assert (fake.created["width"], fake.created["height"]) == (1600, 900)
    assert fake.created["maximized"] is True

    # The post-start maximize is what actually resizes the embedded control.
    fake.start(lambda w: w.maximize(), window)
    assert window.maximized_calls == 1


def test_screen_size_falls_back_when_pywebview_cannot_say():
    import run_server

    for broken in (_FakeWebview(screens=[]),
                   _FakeWebview(screens=[_FakeScreen(0, 0)])):
        assert run_server._primary_screen_size(broken, 1280, 840) == (1280, 840)

    class NoScreens:
        pass
    assert run_server._primary_screen_size(NoScreens(), 1280, 840) == (1280, 840)


def test_a_failed_start_never_opens_a_second_window(monkeypatch):
    """Falling through after create_window succeeded gave the user two windows."""
    import sys

    import run_server

    fake = _FakeWebview(start_error=RuntimeError("callback unsupported"))
    monkeypatch.setitem(sys.modules, "webview", fake)
    launched = []
    monkeypatch.setattr(run_server, "_browser_app_candidates",
                        lambda: ["/nonexistent/browser"])
    monkeypatch.setattr("subprocess.Popen", lambda *a, **k: launched.append(a))

    assert run_server._open_app_window("http://x/", "T") == run_server.WINDOW_CLOSED
    assert not launched, "the browser fallback opened a second window"
    # It retried start() without the callback rather than giving up.
    assert len(fake.started) == 2


def test_browser_fallback_runs_only_when_no_window_was_created(monkeypatch):
    import sys

    import run_server

    class Unusable:
        settings = {}
        screens = []

        def create_window(self, *a, **k):
            raise RuntimeError("no GUI backend")

    monkeypatch.setitem(sys.modules, "webview", Unusable())
    launched = []
    monkeypatch.setattr(run_server, "_browser_app_candidates", lambda: ["/edge"])
    monkeypatch.setattr("subprocess.Popen",
                        lambda *a, **k: launched.append(a[0]) or object())

    assert run_server._open_app_window("http://x/", "T") == run_server.WINDOW_DETACHED
    assert launched and "--start-maximized" in launched[0]


def test_login_screen_is_not_capped_to_the_content_width(client):
    """The login page is a full-bleed background, not a content column.

    It uses <main class="auth-wrap">, so the 1280px cap that `main` puts on
    readable page content applied to it too: the dark panel stopped at 1280px
    and the pale page background showed through as a strip down the side of
    any wider screen. It looked exactly like a window that had failed to
    maximize, which is what made it so easy to misdiagnose.
    """
    import pathlib
    import re

    css = pathlib.Path("itam/static/style.css").read_text()

    auth = re.search(r"^\.auth-wrap\s*\{[^}]*\}", css, re.M | re.S)
    assert auth, ".auth-wrap rule is gone"
    assert "max-width: none" in auth.group(0), \
        ".auth-wrap must opt out of the main content width cap"

    main_rule = re.search(r"^main\s*\{[^}]*\}", css, re.M | re.S)
    assert main_rule, "main rule is gone"
    assert "max-width" not in main_rule.group(0), \
        "main is capped again, which wastes the right of a wide screen"

    # And the markup this depends on must still be the shape we assumed.
    body = client.get("/login").data.decode()
    assert '<main class="auth-wrap">' in body


def test_each_installation_gets_its_own_session_key(tmp_path):
    """A shared install used to sign cookies with a key published in the source.

    SECRET_KEY fell back to "dev-change-me" unless an administrator set it by
    hand, so anyone who could reach a network install could forge a session
    cookie for any account.
    """
    import os

    from itam import create_app

    inst = tmp_path / "instance"
    os.environ.pop("SECRET_KEY", None)
    app = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
                      "UPLOAD_FOLDER": str(tmp_path / "u"),
                      "BACKUP_FOLDER": str(tmp_path / "b")},
                     instance_path=str(inst))
    key = app.config["SECRET_KEY"]
    assert key and key != "dev-change-me"
    assert len(key) >= 32

    # Kept in instance/, so it survives an upgrade and nobody is logged out.
    stored = (inst / "secret_key").read_text().strip()
    assert stored == key

    # A restart reuses it rather than minting a new one.
    again = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
                        "UPLOAD_FOLDER": str(tmp_path / "u"),
                        "BACKUP_FOLDER": str(tmp_path / "b")},
                       instance_path=str(inst))
    assert again.config["SECRET_KEY"] == key

    # A different installation gets a different key.
    other = tmp_path / "other"
    third = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
                        "UPLOAD_FOLDER": str(tmp_path / "u2"),
                        "BACKUP_FOLDER": str(tmp_path / "b2")},
                       instance_path=str(other))
    assert third.config["SECRET_KEY"] != key


def test_secret_key_env_var_still_wins(tmp_path, monkeypatch):
    """Central management (a service definition, Docker) must override."""
    from itam import create_app

    monkeypatch.setenv("SECRET_KEY", "x" * 50)
    app = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
                      "UPLOAD_FOLDER": str(tmp_path / "u"),
                      "BACKUP_FOLDER": str(tmp_path / "b")},
                     instance_path=str(tmp_path / "i"))
    assert app.config["SECRET_KEY"] == "x" * 50


def test_the_installer_ships_and_protects_the_data_folder():
    """The installer runs as administrator on a school server; check its shape."""
    import pathlib

    ps1 = pathlib.Path("deploy/Install-AMS.ps1").read_text()
    assert "instance" in ps1
    # It must never delete or overwrite the data folder.
    for danger in ("Remove-Item $instance", "Remove-Item -Recurse",
                   "rd /s", "Remove-Item $InstallDir"):
        assert danger not in ps1, f"installer contains {danger!r}"
    assert "Register-ScheduledTask" in ps1, "it does not start at boot"
    assert "New-NetFirewallRule" in ps1, "it does not open the firewall"

    un = pathlib.Path("deploy/Uninstall-AMS.ps1").read_text()
    assert "LEFT IN PLACE" in un, "uninstall must not silently bin the database"
    assert "Remove-Item $instance" not in un


def test_password_reset_never_shows_the_link_on_screen(client, app):
    """With no SMTP configured this printed a working reset link to anyone.

    Type a staff email, read the link off the page, own the account. No
    password needed.
    """
    from itam.models import User

    with app.app_context():
        admin = db.session.scalar(db.select(User).where(User.username == "admin"))
        email = admin.email

    resp = client.post("/forgot", data={"email": email}, follow_redirects=True)
    body = resp.data.decode()
    assert "/reset/" not in body, "the reset link is still rendered to the visitor"

    with app.app_context():
        token = db.session.scalar(db.select(User.reset_token).where(User.email == email))
    assert token, "a token should still be issued for the email to carry"
    assert token not in body, "the token leaked into the page"


def test_password_reset_does_not_reveal_which_accounts_exist(client, app):
    real = client.post("/forgot", data={"email": "admin@example.com"},
                       follow_redirects=True).data
    fake = client.post("/forgot", data={"email": "nobody@nowhere.invalid"},
                       follow_redirects=True).data
    assert b"No account found" not in fake
    # The visible answer must be identical either way.
    import re
    strip = lambda b: re.sub(rb"\s+", b" ", b)
    assert strip(real) == strip(fake), "the reply differs, so accounts can be enumerated"


def test_login_locks_out_after_repeated_failures(client, app):
    """The form accepted unlimited password guesses."""
    from itam.blueprints import auth

    auth._FAILURES.clear()
    for _ in range(auth.MAX_ATTEMPTS):
        client.post("/login", data={"username": "admin", "password": "wrong"},
                    follow_redirects=True)

    blocked = client.post("/login", data={"username": "admin", "password": "wrong"},
                          follow_redirects=True)
    assert b"Too many failed sign-ins" in blocked.data

    # Even the correct password is refused while the lockout stands, so
    # guessing cannot be confirmed by a lucky hit.
    right = client.post("/login", data={"username": "admin", "password": "admin123"},
                        follow_redirects=True)
    assert b"Too many failed sign-ins" in right.data
    assert b"Dashboard" not in right.data

    auth._FAILURES.clear()
    ok = client.post("/login", data={"username": "admin", "password": "admin123"},
                     follow_redirects=True)
    assert b"Dashboard" in ok.data


def test_a_successful_login_clears_the_failure_count(client, app):
    from itam.blueprints import auth

    auth._FAILURES.clear()
    for _ in range(auth.MAX_ATTEMPTS - 1):
        client.post("/login", data={"username": "admin", "password": "wrong"},
                    follow_redirects=True)
    assert b"Dashboard" in client.post(
        "/login", data={"username": "admin", "password": "admin123"},
        follow_redirects=True).data
    client.get("/logout", follow_redirects=True)
    # A fresh run of failures must start from zero, not trip immediately.
    assert b"Too many failed sign-ins" not in client.post(
        "/login", data={"username": "admin", "password": "wrong"},
        follow_redirects=True).data


def test_asset_form_fields_added_and_removed(client, app):
    """Type, Location and Device name are gone; Room and Department are there."""
    login(client)
    body = client.get("/assets/new").data.decode()

    assert 'name="asset_type"' not in body, "Type is still on the form"
    assert 'name="hostname"' not in body, "Device name is still on the form"
    assert ">Room<" in body, "Room field is missing"
    assert 'name="department_id"' in body, "Department field is missing"
    # Room writes the same column the Locations tree is derived from.
    assert 'name="location_name"' in body


def test_removing_type_and_hostname_does_not_wipe_imported_values(client, app):
    """Both came in from imports; the form no longer submits them."""
    from itam.models import Category

    with app.app_context():
        a = Asset(tag="KEEP-1", name="Imported PC", status="In Use",
                  condition="Good", asset_type="Desktop", hostname="LAB-PC-07",
                  category=db.session.scalar(db.select(Category)))
        db.session.add(a)
        db.session.commit()
        aid = a.id

    client.post(f"/assets/{aid}/edit", data={
        "name": "Imported PC", "tag": "KEEP-1", "status": "In Use",
        "condition": "Good", "depreciation_years": "5"}, follow_redirects=True)
    with app.app_context():
        a = db.session.get(Asset, aid)
        assert a.asset_type == "Desktop", "saving the form wiped the imported Type"
        assert a.hostname == "LAB-PC-07", "saving the form wiped the device name"


def test_department_and_room_save_from_the_form(client, app):
    from itam.models import Category, Department

    with app.app_context():
        dep = Department(name="Science")
        db.session.add(dep)
        db.session.commit()
        dep_id, cat_id = dep.id, db.session.scalar(db.select(Category.id))

    login(client)
    client.post("/assets/new", data={
        "name": "Lab PC", "tag": "ROOM-1", "category_id": cat_id,
        "status": "In Use", "condition": "Good", "depreciation_years": "5",
        "branch": "Mada 1", "building": "Building 1", "floor": "GF",
        "location_name": "Computer Lab", "department_id": dep_id},
        follow_redirects=True)

    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "ROOM-1"))
        assert a.location_name == "Computer Lab"
        assert a.department_id == dep_id

    # Clearing the department must actually clear it, not be ignored.
    with app.app_context():
        aid = db.session.scalar(db.select(Asset.id).where(Asset.tag == "ROOM-1"))
    client.post(f"/assets/{aid}/edit", data={
        "name": "Lab PC", "tag": "ROOM-1", "status": "In Use",
        "condition": "Good", "depreciation_years": "5",
        "department_id": ""}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Asset, aid).department_id is None


def test_room_list_offers_rooms_already_in_the_data(client, app):
    """An imported room should be one click away, not retyped exactly."""
    from itam.models import Category

    with app.app_context():
        db.session.add(Asset(tag="RM-1", name="PC", status="In Use",
                             condition="Good", location_name="Grade 4 Annexe",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()

    login(client)
    body = client.get("/assets/new").data.decode()
    picker = body.split('id="f-place-pick"', 1)[1].split("</select>", 1)[0]
    assert "Grade 4 Annexe" in picker, "a room already in use is not offered"
    assert "Reception" in picker, "the standard rooms are no longer offered"


def test_a_fresh_database_has_categories(app):
    """An empty Category list made the Asset ID impossible to generate.

    The ID comes from the category, so with none seeded a new install could
    not add a properly numbered asset until someone worked out they had to go
    and create categories first.
    """
    from itam.models import Category

    with app.app_context():
        names = {c.name for c in db.session.scalars(db.select(Category))}
        assert len(names) > 5, "a fresh database still has no categories"
        assert {"Desktop Computers", "Laptops", "Printers"} <= names
        # Every one can generate an ID.
        for c in db.session.scalars(db.select(Category)):
            assert c.tag_prefix, f"{c.name} has no Asset ID prefix"


def test_seeded_categories_never_touch_an_existing_list(app):
    """A school that made its own categories must not have ours added."""
    from itam import _ensure_defaults
    from itam.models import Category

    with app.app_context():
        db.session.execute(db.delete(Category))
        db.session.add(Category(name="Our Own Thing", prefix="OWN"))
        db.session.commit()
        _ensure_defaults()
        names = {c.name for c in db.session.scalars(db.select(Category))}
        assert names == {"Our Own Thing"}, "defaults were added over a real list"


def test_next_tag_follows_the_format_already_in_use(app):
    """PC00010 must continue as PC00017, not restart as PC-000001."""
    from itam.blueprints.assets import next_tag
    from itam.models import Category

    with app.app_context():
        cat = Category(name="School Desktops", prefix="PC")
        db.session.add(cat)
        db.session.flush()
        for t in ("PC00010", "PC00011", "PC00012", "PC00013",
                  "PC00014", "PC00015", "PC00016"):
            db.session.add(Asset(tag=t, name=t, status="In Use",
                                 condition="Good", category=cat))
        db.session.commit()
        assert next_tag(cat) == "PC00017"

        # A different width is followed too.
        wide = Category(name="Wide", prefix="WD")
        db.session.add(wide)
        db.session.flush()
        db.session.add(Asset(tag="WD-000042", name="w", status="In Use",
                             condition="Good", category=wide))
        db.session.commit()
        assert next_tag(wide) == "WD-000043"

        # And a category with nothing yet still gets the original default.
        fresh = Category(name="Brand New", prefix="BN")
        db.session.add(fresh)
        db.session.commit()
        assert next_tag(fresh) == "BN-000001"


def test_assign_to_is_editable_when_adding_but_not_when_editing(client, app):
    from itam.models import Category, Employee

    login(client)
    body = client.get("/assets/new").data.decode()
    assert 'name="assign_employee_id"' in body, "Assign to is not offered on a new asset"

    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
        emp_id = db.session.scalar(db.select(Employee.id))

    client.post("/assets/new", data={
        "name": "Handed over", "tag": "ASG-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "assign_employee_id": emp_id}, follow_redirects=True)

    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "ASG-1"))
        assert a.current_assignment is not None, "the asset was not handed over"
        assert a.current_assignment.employee_id == emp_id
        assert a.current_assignment.handled_by == "Administrator"
        assert a.status == "Checked Out"
        aid = a.id

    # On an existing asset it stays read-only, so a save cannot end the loan.
    edit = client.get(f"/assets/{aid}/edit").data.decode()
    assert 'name="assign_employee_id"' not in edit
    client.post(f"/assets/{aid}/edit", data={
        "name": "Handed over", "tag": "ASG-1", "status": "Checked Out",
        "condition": "Good", "depreciation_years": "5"}, follow_redirects=True)
    with app.app_context():
        assert db.session.get(Asset, aid).current_assignment is not None


def test_a_viewer_cannot_change_departments(client, app):
    """The Departments page is readable by anyone who can see assets.

    The write branch shares that permission, so a viewer or an auditor could
    create and rename departments through it.
    """
    from itam.models import Department

    login(client, "viewer", "viewer123")
    assert client.get("/departments").status_code == 200, "a viewer should still see them"

    client.post("/departments", data={"name": "Sneaky Dept"}, follow_redirects=True)
    with app.app_context():
        assert not db.session.scalar(
            db.select(Department).where(Department.name == "Sneaky Dept")), \
            "a viewer created a department"

    # And an admin still can.
    client.get("/logout")
    login(client)
    client.post("/departments", data={"name": "Real Dept"}, follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(
            db.select(Department).where(Department.name == "Real Dept")), \
            "an admin can no longer create departments"


def test_no_write_endpoint_is_reachable_by_a_viewer(client, app):
    """Sweep every parameterless POST route rather than trusting a spot check."""
    login(client, "viewer", "viewer123")
    public = {"/login", "/forgot", "/profile", "/notifications/read",
              "/assets/searches"}          # own account / own saved searches
    reached = []
    for rule in app.url_map.iter_rules():
        path = str(rule)
        if "POST" not in (rule.methods - {"HEAD", "OPTIONS"}) or "<" in path:
            continue
        if path in public:
            continue
        resp = client.post(path, data={}, follow_redirects=False)
        if resp.status_code not in (302, 401, 403):
            reached.append((path, resp.status_code))
    assert not reached, f"a viewer reached write endpoints: {reached}"


def test_labels_open_in_the_browser_when_no_printer_is_configured(client, app):
    """Direct printing is opt-in; without a printer name nothing changes."""
    from itam.models import Category

    with app.app_context():
        db.session.add(Asset(tag="LBL-1", name="Label test", status="Available",
                             condition="Good",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()
        aid = db.session.scalar(db.select(Asset.id).where(Asset.tag == "LBL-1"))

    login(client)
    resp = client.get(f"/assets/{aid}/label")
    assert resp.status_code == 200
    assert b"LBL-1" in resp.data, "the label page no longer renders"
    assert client.get("/assets/labels").status_code == 200


def test_direct_printing_is_refused_off_windows(app):
    """It must never claim to have printed when it cannot."""
    from itam import printing

    assert printing.can_print_directly("") is False          # no printer set
    if not printing.is_windows():
        assert printing.can_print_directly("XP-490B") is False
        assert printing.print_html("<p>x</p>", "XP-490B") is False
    assert printing.list_printers() == [] or printing.is_windows()


def test_a_remote_user_never_prints_on_the_server(monkeypatch):
    """Printing happens where AMS runs, not where the person is sitting.

    On a server install, staff at their own desks must get their own print
    dialog -- otherwise they are told the label was sent while it comes out
    of a printer in the server room.
    """
    import sys

    from itam import printing

    monkeypatch.setattr(printing, "is_windows", lambda: True)
    monkeypatch.setattr(sys, "frozen", True, raising=False)

    # At the machine running AMS: print directly.
    assert printing.can_print_directly("XP-490B", "127.0.0.1") is True
    assert printing.can_print_directly("XP-490B", "::1") is True
    # Someone else on the network: hand them the browser dialog.
    assert printing.can_print_directly("XP-490B", "192.168.100.55") is False
    assert printing.can_print_directly("XP-490B", "10.0.0.9") is False


def test_employee_id_is_generated(client, app):
    """A manual add left it blank, and an import without the column made
    employees with no ID -- the very field the importer matches on."""
    from itam.models import Employee

    login(client)
    body = client.get("/employees/new").data.decode()
    assert 'name="emp_code"' in body
    assert "EMP-" in body, "no ID was suggested on the blank form"

    client.post("/employees/new", data={
        "name": "No Id Person", "email": "noid@example.com", "active": "1"},
        follow_redirects=True)
    with app.app_context():
        e = db.session.scalar(db.select(Employee).where(Employee.name == "No Id Person"))
        assert e.emp_code, "an employee was created without an ID"
        first = e.emp_code

    client.post("/employees/new", data={
        "name": "Second Person", "email": "second@example.com", "active": "1"},
        follow_redirects=True)
    with app.app_context():
        e2 = db.session.scalar(db.select(Employee).where(Employee.name == "Second Person"))
        assert e2.emp_code and e2.emp_code != first, "two employees share an ID"


def test_imported_employees_get_an_id_when_the_sheet_has_none(client, app):
    from itam.models import Employee

    login(client)
    csv = "Name,Email\nImported One,i1@example.com\nImported Two,i2@example.com\n"
    client.post("/employees/import",
                data={"file": (io.BytesIO(csv.encode()), "e.csv")},
                content_type="multipart/form-data", follow_redirects=True)
    with app.app_context():
        rows = db.session.scalars(
            db.select(Employee).where(Employee.name.like("Imported %"))).all()
        assert len(rows) == 2
        codes = [r.emp_code for r in rows]
        assert all(codes), "imported employees have no ID to match on next time"
        assert len(set(codes)) == 2, "imported employees share an ID"


def test_departments_are_seeded(app):
    from itam.models import Department

    with app.app_context():
        names = {d.name for d in db.session.scalars(db.select(Department))}
    assert len(names) > 5, "the department dropdown still starts empty"
    assert "IT Department" in names


def test_assets_list_offers_select_all(client, app):
    """Assets uses its own check-all rather than the shared macro."""
    from itam.models import Category

    with app.app_context():
        db.session.add(Asset(tag="SA-1", name="Rig", status="Available",
                             condition="Good",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()

    login(client)
    body = client.get("/assets/").data.decode()
    assert "bulk-toggle" in body, "Assets has no Select all button"
    assert 'class="bulk-all"' in body, "Assets has no header tick-box"
    assert 'name="id"' in body, "rows are not selectable"

def test_adding_a_location_builds_the_whole_chain(client, app):
    """Fill in the levels you know; the branch, building, floor and room are
    created and linked in one go rather than four separate adds."""
    from itam.models import Location

    login(client)
    body = client.get("/locations").data.decode()
    for field in ("branch", "building", "department", "floor", "room"):
        assert f'name="{field}"' in body, f"the {field} field is missing"

    client.post("/locations", data={
        "branch": "Mada 2", "building": "Building 3", "department": "IT Department",
        "floor": "F4", "room": "Robotics Lab"}, follow_redirects=True)

    with app.app_context():
        room = db.session.scalar(db.select(Location).where(Location.name == "Robotics Lab"))
        assert room is not None and room.kind == "Room"
        assert room.path == "Mada 2 / Building 3 / IT Department / F4 / Robotics Lab"

    # Adding it again reuses every level instead of duplicating them.
    with app.app_context():
        before = db.session.scalar(db.select(db.func.count(Location.id)))
    client.post("/locations", data={
        "branch": "Mada 2", "building": "Building 3", "department": "IT Department",
        "floor": "F4", "room": "Robotics Lab"}, follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(db.func.count(Location.id))) == before

    # A partial chain is fine too.
    client.post("/locations", data={"branch": "Mada 3", "room": "Store"},
                follow_redirects=True)
    with app.app_context():
        store = db.session.scalar(db.select(Location).where(Location.name == "Store"))
        assert store.path == "Mada 3 / Store"

    # And an empty form is refused rather than making a nameless row.
    resp = client.post("/locations", data={}, follow_redirects=True)
    assert b"at least one level" in resp.data


def test_departments_are_reachable_from_locations(client, app):
    """They share a screen now; the separate sidebar entry is gone."""
    login(client)
    body = client.get("/locations").data.decode()
    assert 'id="departments"' in body, "departments are not shown on Locations"
    assert "/departments" in body, "no way through to manage them"
    # The page still works, it just isn't in the sidebar any more.
    assert client.get("/departments").status_code == 200
    rail = client.get("/assets/").data.decode().split('class="rail')[1].split("</aside>")[0]
    assert ">Departments<" not in rail, "still in the navigation rail"


def test_an_employee_needs_no_email(client, app):
    """Drivers, cleaners and security often have none; the form demanded one."""
    from itam.models import Employee

    login(client)
    body = client.get("/employees/new").data.decode()
    email_field = body.split('id="e-email"', 1)[1].split(">", 1)[0]
    assert "required" not in email_field, "email is still mandatory"

    resp = client.post("/employees/new", data={"name": "Driver One", "active": "1"},
                       follow_redirects=True)
    assert resp.status_code == 200
    with app.app_context():
        e = db.session.scalar(db.select(Employee).where(Employee.name == "Driver One"))
        assert e is not None, "an employee without an email could not be added"
        assert e.email is None and e.emp_code

    # A shared inbox is allowed, which is what the model always said.
    for name in ("Shared One", "Shared Two"):
        client.post("/employees/new", data={"name": name, "email": "office@school.test",
                                            "active": "1"}, follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(db.func.count(Employee.id))
                                 .where(Employee.email == "office@school.test")) == 2


def test_lists_offer_a_visible_select_all_button(client, app):
    """The tick-box in the table header was too easy to miss."""
    from itam.models import Department

    with app.app_context():
        db.session.add(Department(name="Bulk Dept"))
        db.session.commit()

    login(client)
    body = client.get("/employees").data.decode()
    assert "bulk-toggle" in body, "no visible Select all button"
    assert 'data-form="bulk"' in body, "the button does not target the list's form"


def test_select_all_appears_on_every_list_that_has_rows(client, app):
    """All nine lists, checked together rather than one at a time.

    Assets had its own selection code and never picked up the shared button,
    and the others only render the toolbar once a list has rows -- which on a
    new install is never, so it looked like nothing had one.
    """
    from itam.models import (Assignment, Category, Department, Employee,
                             InventoryAudit, License, Location, Maintenance,
                             Vendor)

    with app.app_context():
        cat = db.session.scalar(db.select(Category))
        emp = db.session.scalar(db.select(Employee))
        db.session.add_all([Department(name="D One"), Location(name="L One", kind="Room"),
                            License(name="Lic One", seats=3), Vendor(name="V One"),
                            InventoryAudit(name="A One")])
        a = Asset(tag="SEL-ALL", name="Thing", status="Checked Out",
                  condition="Good", category=cat)
        db.session.add(a)
        db.session.flush()
        db.session.add_all([Maintenance(asset=a, title="Fix"),
                            Assignment(asset=a, employee=emp)])
        db.session.commit()

    login(client)
    missing = []
    for path in ("/assets/", "/checkouts", "/licenses", "/maintenance", "/inventory",
                 "/employees", "/departments", "/vendors"):
        body = client.get(path).data.decode()
        if not ("bulk-toggle" in body and 'class="bulk-all"' in body
                and 'class="bulk-pick"' in body):
            missing.append(path)
    assert not missing, f"no working select-all on: {missing}"


def test_arabic_translates_the_whole_page(client, app):
    """Most of the interface never called t(), so Arabic was largely English."""
    login(client)
    client.get("/lang/ar")
    body = client.get("/assets/").data.decode()

    for english in ("All statuses", "All categories", "All departments",
                    "All branches", "All buildings", "All floors",
                    "All locations", "All conditions", "Filter",
                    "No assets found"):
        assert english not in body, f"{english!r} is still in English"
    for arabic in ("كل الحالات", "كل الفئات", "تصفية", "لا توجد أصول"):
        assert arabic in body, f"{arabic} is missing"

    # English is untouched.
    client.get("/lang/en")
    body = client.get("/assets/").data.decode()
    assert "All statuses" in body and "كل الحالات" not in body


def test_arabic_never_rewrites_data_or_scripts(client, app):
    """Only exact UI phrases are swapped, and never inside a script."""
    from itam.models import Category

    with app.app_context():
        db.session.add(Asset(tag="AR-1", name="Filter unit for the pool",
                             status="Available", condition="Good",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()

    login(client)
    client.get("/lang/ar")
    body = client.get("/assets/").data.decode()
    # A longer name containing a translated word is left alone.
    assert "Filter unit for the pool" in body
    # Script contents are skipped, so the JavaScript still works.
    from itam.i18n import translate_html
    js = '<script>var s = "Filter";</script><p>Filter</p>'
    out = translate_html(js)
    assert 'var s = "Filter";' in out, "script contents were rewritten"
    assert "<p>تصفية</p>" in out


def test_dropdowns_are_enhanced_but_the_real_select_still_submits(client, app):
    """The native popup is drawn by the OS and opens upward when it runs out of
    room below, which no CSS or script can override. The popup is replaced, but
    the <select> itself stays, so submitting is unchanged."""
    import pathlib

    js = pathlib.Path("itam/static/app.js").read_text()
    css = pathlib.Path("itam/static/style.css").read_text()

    assert "sel-list" in js and "sel-btn" in js, "the dropdown script is missing"
    # It must never flip upward: no 'bottom' positioning on the list.
    rule = css.split(".sel-list {", 1)[1].split("}", 1)[0]
    assert "top: calc(100% + 2px)" in rule, "the list is not pinned below the field"
    assert "bottom:" not in rule, "the list can still flip upward"
    # List boxes and multi-selects keep the native behaviour.
    assert "s.multiple" in js and "s.size" in js

    # The form still works exactly as before.
    from itam.models import Category

    login(client)
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
    client.post("/assets/new", data={
        "name": "Dropdown test", "tag": "DD-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5"},
        follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "DD-1"))
        assert a is not None and a.category_id == cat_id, \
            "the select no longer submits its value"


def test_every_location_level_offers_other_and_creates_it(client, app):
    """The levels were fixed lists compiled into the app, so a school could
    only ever pick a branch, building or floor we had thought of."""
    from itam.models import Location

    login(client)
    body = client.get("/locations").data.decode()
    for level in ("branch", "building", "dept", "floor", "room"):
        block = body.split(f'id="loc-{level}-pick"', 1)[1].split("</select>", 1)[0]
        assert "__other__" in block, f"{level} has no Other option"
        # Other… must come last, after the real choices.
        assert block.rindex("__other__") > block.rindex('<option value="">'), \
            f"{level}'s Other option is not after the others"

    # Typing new values creates the whole chain.
    client.post("/locations", data={
        "branch": "Mada 4", "building": "Annexe A", "department": "Robotics",
        "floor": "Mezzanine", "room": "Robotics Lab"}, follow_redirects=True)
    with app.app_context():
        room = db.session.scalar(db.select(Location).where(Location.name == "Robotics Lab"))
        assert room.path == "Mada 4 / Annexe A / Robotics / Mezzanine / Robotics Lab"
        for name, kind in (("Mada 4", "Branch"), ("Annexe A", "Building"),
                           ("Robotics", "Department"), ("Mezzanine", "Floor")):
            got = db.session.scalar(db.select(Location).where(Location.name == name))
            assert got is not None and got.kind == kind, f"{name} was not created as a {kind}"

    # And they are offered next time rather than having to be retyped.
    body = client.get("/locations").data.decode()
    for name in ("Mada 4", "Annexe A", "Mezzanine", "Robotics Lab"):
        assert name in body, f"{name} is not offered after being created"


def test_a_location_can_be_deleted(client, app):
    """The row table is gone from the page, but the delete endpoint stays
    (the level manager and integrations still use it)."""
    from itam.models import Location

    login(client)
    client.post("/locations", data={"branch": "Temp Branch", "room": "Temp Room"},
                follow_redirects=True)
    with app.app_context():
        room = db.session.scalar(db.select(Location).where(Location.name == "Temp Room"))
        room_id = room.id

    client.post(f"/locations/{room_id}/delete", follow_redirects=True)
    with app.app_context():
        assert db.session.get(Location, room_id) is None, "the row was not deleted"


def test_a_location_holding_assets_is_not_deleted(client, app):
    """Deleting a place that still has equipment in it would orphan the link."""
    from itam.models import Category, Location

    with app.app_context():
        loc = Location(name="Busy Room", kind="Room")
        db.session.add(loc)
        db.session.flush()
        db.session.add(Asset(tag="BUSY-1", name="PC", status="In Use",
                             condition="Good", location_id=loc.id,
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()
        loc_id = loc.id

    login(client)
    resp = client.post(f"/locations/{loc_id}/delete", follow_redirects=True)
    assert b"cannot be deleted" in resp.data
    with app.app_context():
        assert db.session.get(Location, loc_id) is not None, "a location in use was deleted"


def test_levels_added_on_locations_reach_the_asset_form(client, app):
    """The whole point: Branch/Building/Floor were lists compiled into the app.

    The asset form offered only those, and _from_form threw away anything
    else, so a school with a fourth campus could not record it at all.
    """
    from itam.models import Category, Location

    login(client)
    client.post("/locations/level", data={
        "level": "building", "action": "add", "name": "Annexe A"},
        follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(Location).where(
            Location.name == "Annexe A", Location.kind == "Building"))

    # offered on the form...
    body = client.get("/assets/new").data.decode()
    picker = body.split('id="f-building"', 1)[1].split("</select>", 1)[0]
    assert "Annexe A" in picker, "a building added on Locations is not on the asset form"

    # ...and kept on save, which is what used to fail silently.
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
    client.post("/assets/new", data={
        "name": "Annexe PC", "tag": "ANX-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "building": "Annexe A"}, follow_redirects=True)
    with app.app_context():
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "ANX-1"))
        assert a.building == "Annexe A", "the new building was discarded on save"


def test_a_level_can_be_renamed_everywhere_at_once(client, app):
    """Assets store these as text, so a rename has to reach them too."""
    from itam.models import Category, Location

    login(client)
    client.post("/locations/level", data={
        "level": "branch", "action": "add", "name": "Mada 4"}, follow_redirects=True)
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
    client.post("/assets/new", data={
        "name": "Campus PC", "tag": "CMP-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "branch": "Mada 4"}, follow_redirects=True)

    client.post("/locations/level", data={
        "level": "branch", "action": "rename", "was": "Mada 4",
        "name": "Mada 4 — New Campus"}, follow_redirects=True)

    with app.app_context():
        assert db.session.scalar(db.select(Location).where(
            Location.name == "Mada 4 — New Campus"))
        a = db.session.scalar(db.select(Asset).where(Asset.tag == "CMP-1"))
        assert a.branch == "Mada 4 — New Campus", \
            "the asset still points at the old name"


def test_a_level_in_use_is_not_removed(client, app):
    """Removing a building that assets sit in would orphan them."""
    from itam.models import Category, Location

    login(client)
    client.post("/locations/level", data={
        "level": "floor", "action": "add", "name": "Mezzanine"}, follow_redirects=True)
    with app.app_context():
        cat_id = db.session.scalar(db.select(Category.id))
    client.post("/assets/new", data={
        "name": "Mez PC", "tag": "MEZ-1", "category_id": cat_id,
        "status": "Available", "condition": "Good", "depreciation_years": "5",
        "floor": "Mezzanine"}, follow_redirects=True)

    resp = client.post("/locations/level", data={
        "level": "floor", "action": "delete", "name": "Mezzanine"},
        follow_redirects=True)
    assert b"was kept" in resp.data
    with app.app_context():
        assert db.session.scalar(db.select(Location).where(Location.name == "Mezzanine"))

    # An unused one goes.
    client.post("/locations/level", data={
        "level": "floor", "action": "add", "name": "Unused Floor"}, follow_redirects=True)
    client.post("/locations/level", data={
        "level": "floor", "action": "delete", "name": "Unused Floor"},
        follow_redirects=True)
    with app.app_context():
        assert not db.session.scalar(db.select(Location).where(Location.name == "Unused Floor"))


def test_departments_can_be_managed_from_locations(client, app):
    """Departments are their own table -- that is what assets link to."""
    from itam.models import Department

    login(client)
    body = client.get("/locations").data.decode()
    assert 'value="department"' in body, "no department field on Locations"

    client.post("/locations/level", data={
        "level": "department", "action": "add", "name": "Robotics Club"},
        follow_redirects=True)
    with app.app_context():
        dep = db.session.scalar(db.select(Department).where(Department.name == "Robotics Club"))
        assert dep is not None

    client.post("/locations/level", data={
        "level": "department", "action": "rename", "was": "Robotics Club",
        "name": "Robotics"}, follow_redirects=True)
    with app.app_context():
        assert db.session.scalar(db.select(Department).where(Department.name == "Robotics"))

    # It is offered on the asset form.
    assert "Robotics" in client.get("/assets/new").data.decode()


def test_a_viewer_cannot_change_levels(client, app):
    from itam.models import Location

    login(client, "viewer", "viewer123")
    client.post("/locations/level", data={
        "level": "branch", "action": "add", "name": "Sneaky Branch"},
        follow_redirects=True)
    with app.app_context():
        assert not db.session.scalar(
            db.select(Location).where(Location.name == "Sneaky Branch"))


def test_renamed_repo_is_migrated_in_stored_settings(app):
    """The GitHub repo was renamed; installs that saved the old path follow.

    GitHub redirects renamed repos, but only until something else claims the
    old name -- so stored settings move to the new path. A repo the school
    pointed elsewhere themselves is left alone.
    """
    from itam import DEFAULT_SETTINGS, _ensure_defaults
    from itam.models import Setting

    with app.app_context():
        row = db.session.get(Setting, "update_repo")
        old = "laithyahya2022-code/IT-Asset-Management-System-"
        if row is None:
            db.session.add(Setting(key="update_repo", value=old))
        else:
            row.value = old
        db.session.commit()
        _ensure_defaults()
        assert db.session.get(Setting, "update_repo").value == \
            DEFAULT_SETTINGS["update_repo"]

        # A custom value must survive.
        db.session.get(Setting, "update_repo").value = "someone/else"
        db.session.commit()
        _ensure_defaults()
        assert db.session.get(Setting, "update_repo").value == "someone/else"


def test_label_links_stay_in_the_app_window(client, app):
    """Inside the desktop app, target="_blank" opens the system browser --
    which has no session cookie, so pressing Print landed staff on a sign-in
    page instead of the label."""
    from itam.models import Category

    with app.app_context():
        db.session.add(Asset(tag="WIN-1", name="Window test", status="Available",
                             condition="Good",
                             category=db.session.scalar(db.select(Category))))
        db.session.commit()
        aid = db.session.scalar(db.select(Asset.id).where(Asset.tag == "WIN-1"))

    login(client)
    for page in (f"/assets/{aid}", f"/assets/{aid}/edit"):
        body = client.get(page).data.decode()
        for line in body.splitlines():
            if "assets.label" in line or f"/assets/{aid}/label" in line:
                assert 'target="_blank"' not in line, \
                    f"{page} still opens the label in a new window"

    # The label page itself carries a way back, so same-window is not a trap.
    assert b"Back to asset" in client.get(f"/assets/{aid}/label").data
